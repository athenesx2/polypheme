##################################
# polypheme.py                   #
# Par Eytan Benharrous           #
#                                #
#                  /^\--/^\      #
#                  | O  O |      #
#    Whoo!!        \  \/  /      #
#                  / ¤¤  / \     #
#      Whoo!!     |¤¤¤¤¤|;;;|    #
#                 |¤¤¤¤¤|;;;|    #
#                 | ¤¤¤¤ \;;|    #
#                  \ ¤¤    \|    #
####################(((##(((######

import tkinter as tk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import ttk
from matplotlib.collections import PolyCollection
from tkinter import filedialog
import openpyxl
import random
import shutil
import matplotlib.ticker as mticker
from PIL import ImageTk, Image, ImageEnhance
from math import log
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import numpy as np
import time as t
import openpyxl.utils.exceptions


class MovableImage(tk.Frame):
    def __init__(self, parent, nom):
        tk.Frame.__init__(self, parent)
        self.parent = parent

        self.start_coords = {"x": 0, "y": 0, "move": False}

        self.start_coords_main = {"x": 0, "y": 0, "move": False}

        self.image = Image.open(nom)

        img = self.image
        enhancer = ImageEnhance.Brightness(img)
        im_output = enhancer.enhance(1)
        self.image2 = im_output
        self.image2.putalpha(250)
        self.main_image = ImageTk.PhotoImage(
            self.image.resize((1000, 1000), Image.ANTIALIAS)
        )
        self.nav_image = ImageTk.PhotoImage(
            self.image2.resize((50, 50), Image.ANTIALIAS)
        )
        self.tempon_image = ImageTk.PhotoImage(
            self.image.resize((200, 200), Image.ANTIALIAS)
        )

        self.main_canvas = tk.Canvas(self, width=200, height=200, highlightthickness=0)
        self.main_canvas.pack()
        self.main_image_id = self.main_canvas.create_image(
            (0, 0), image=self.tempon_image, anchor="nw", tags="main_image"
        )

        self.nav_canvas = tk.Canvas(
            self.main_canvas,
            width=50,
            height=50,
            highlightthickness=1,
            highlightbackground="grey",
        )
        self.main_canvas.create_window(
            (150, 150), window=self.nav_canvas, anchor="nw", tags="nav_canvas"
        )
        self.nav_canvas.create_image((0, 0), image=self.nav_image, anchor="nw")

        self.nav_canvas.bind("<Button-1>", self.set_start_coords)
        self.nav_canvas.bind("<B1-Motion>", self.move_coords)

        self.nav_canvas.bind("<ButtonRelease-1>", self.fini_coords)

    def fini_coords(self, event):
        
        self.nav_canvas.delete(self.nav_box)
        self.main_canvas.itemconfig(self.main_image_id, image=self.tempon_image)
        self.main_canvas.coords(self.main_image_id, 0, 0)

     def set_start_coords(self, event):
        x = event.x - 5
        y = event.y - 5

        if x < 0:
            x = 0
        elif x + 10 > 100:
            x = 40
        if y < 0:
            y = 0
        elif y + 10 > 100:
            y = 40

        self.main_canvas.itemconfig(self.main_image_id, image=self.main_image)

        
        self.nav_box = self.nav_canvas.create_rectangle(
            (x, y, x + 10, y + 10), outline="black"
        )
        self.main_canvas.coords(self.main_image_id, -(x) * 20, -y * 20)

        x1, y1, x2, y2 = self.nav_canvas.coords(self.nav_box)
        if x1 < event.x < x2 and y1 < event.y < y2:
            self.start_coords["x"] = event.x - x1
            self.start_coords["y"] = event.y - y1
            self.start_coords["move"] = True
        else:
            self.start_coords["move"] = False

    def move_coords(self, event):
        if not self.start_coords["move"]:
            return

        dx = self.start_coords["x"]
        dy = self.start_coords["y"]
        x = event.x - dx
        y = event.y - dy

        if x < 0:
            x = 0
        elif x + 10 > 50:
            x = 40
        if y < 0:
            y = 0
        elif y + 10 > 50:
            y = 40

        self.nav_canvas.coords(self.nav_box, x, y, x + 10, y + 10)
        self.main_canvas.coords(self.main_image_id, -(x) * 20, -y * 20)


def lit(fichier, position, nb_octets):
    fichier.seek(position)
    return list(fichier.read(nb_octets))


def lit_entier(fichier, position, nb_octets):
    fichier.seek(position)
    return int.from_bytes(fichier.read(nb_octets), byteorder="little", signed=True)


def ecrit(fichier, position, octet):
    fichier.seek(position)
    fichier.write(bytes([octet]))


def ecrit_liste(fichier, position, octets):
    fichier.seek(position)
    fichier.write(bytes(octets))


def ecrit_entier(fichier, position, entier, nb_octets):
    fichier.seek(position)
    fichier.write(entier.to_bytes(nb_octets, byteorder="little", signed=True))


def initiateur(fichier):
    global position
    ecrit_liste(fichier, position, [0, 255, 255, 0])
    position += 4


def monomerec(fichier, couleur):
    global position
    ecrit_liste(fichier, position, couleur)
    position += 4


def fini_ligne(fichier, tranche):
    global position
    retour = 4 * tranche - (position - 54) % (4 * tranche)
    for i in range(retour // 4):
        ecrit_liste(fichier, position, [255, 255, 255, 0])
        position += 4


def commence_image(fichier):
    global position
    position = 54


def fini_image(fichier, tranche):
    global position
    retour = 4 * tranche**2 - (position - 54) % (4 * tranche**2)
    for i in range(retour // 4):
        ecrit_liste(fichier, position, [255, 255, 255, 0])
        position += 4


def simulmc(
    tranche,
    initiateurs: dict[str : tuple[int, int]],
    reactants: list[tuple[tuple[str, str], int]],
    terminaison: tuple[int, int],
    tstochastic: int,
):
    """entrée: initiateurs:
    -dictionnaire qui associe au nom des reactif et initiateurs
    leurs quantités et vitesse dans un couple:
    {nom 1: (vitesse 1, quantité 1),noinitiateurs[i][1]m 2: (vitesse 2, quantité 2),...]
    -liste de tuple contenant un couple avec le nom des deux reactants,
    et leur constante de vitesse:
    [((nom 1.1, nom 1.2), vitesse 1,concentration de chaine1),((nom 2.1, nom 2.2) vitesse 2,concentration de chaine2),...]
    -tuple contenant les deux vitesse de terminaison.
    -temps stochastique

    retour:
    (temps,liste[tuple(nom du monomere,moyenne de chaine)])
    """
    dicomoyennefabric1 = {}
    dicomoyennefabric2 = {}
    t1 = t.time()
    f = open("simulmc.bmp", "w")
    f.close()
    # crée le fichier
    f = open("simulmc.bmp", "r+b")
    # règle les paramètres de l'image
    ecrit_liste(
        f,
        0,
        [
            66,
            77,
            54,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            54,
            0,
            0,
            0,
            40,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            255,
            255,
            255,
            255,
            1,
            0,
            32,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
        ],
    )
    ecrit_entier(f, 22, -tranche, 4)
    ecrit_entier(f, 18, tranche, 4)
    ecrit_entier(f, 2, 54 + tranche * tranche * 4, 4)

    g = open("simulmc2.bmp", "w")
    g.close()
    # crée le fichier
    g = open("simulmc2.bmp", "r+b")
    # règle les paramètres de l'image
    ecrit_liste(
        g,
        0,
        [
            66,
            77,
            54,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            54,
            0,
            0,
            0,
            40,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            255,
            255,
            255,
            255,
            1,
            0,
            32,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
        ],
    )
    ecrit_entier(g, 22, -tranche, 4)
    ecrit_entier(g, 18, tranche, 4)
    ecrit_entier(g, 2, 54 + tranche * tranche * 4, 4)
    commence_image(f)
    tempo = 0
    dicochaine = {}
    listechaine = 0
    dicocouleur = {}
    chainefini = []
    chainechoisif = []
    couleur = [
        ([0, 0, 255, 0], "#ff0000", (255 / 255, 0, 0)),
        ([87, 201, 0, 0], "#00c957", (0, 201 / 255, 87 / 255)),
        (
            [250, 206, 135, 0],
            "#87cefa",
            (135 / 255, 206 / 255, 250 / 255),
        ),
        (
            [203, 192, 255, 0],
            "#ffc0cb",
            (255 / 255, 192 / 255, 203 / 255),
        ),
        (
            [211, 0, 148, 0],
            "#9400d3",
            (148 / 255, 0, 211 / 255),
        ),
        (
            [0, 85, 204, 0],
            "#cc5500",
            (204 / 255, 85 / 255, 0),
        ),
        (
            [152, 255, 152, 0],
            "#98ff98",
            (152 / 255, 255 / 255, 152 / 255),
        ),
        ([225, 105, 65, 0], "#4169e1", (65 / 255, 105 / 255, 225 / 255)),
        ([255, 0, 255, 0], "#ff00ff", (255 / 255, 0, 255 / 255)),
        ([128, 0, 0, 0], "#000080", (0, 0, 128 / 255)),
        ([128, 0, 128, 0], "#800080", (128 / 255, 0, 128 / 255)),
        ([0, 165, 255, 0], "#ffa500", (255 / 255, 165 / 255, 0)),
        ([50, 205, 50, 0], "#32cd32", (50 / 255, 205 / 255, 50 / 255)),
        ([250, 232, 230, 0], "#e6e8fa", (230 / 255, 232 / 255, 250 / 255)),
        ([95, 17, 224, 0], "#e0115f", (224 / 255, 17 / 255, 95 / 255)),
        ([255, 255, 0, 0], "#00ffff", (0, 255 / 255, 255 / 255)),
        ([30, 105, 210, 0], "#d2691e", (210 / 255, 105 / 255, 30 / 255)),
        ([144, 128, 112, 0], "#708090", (112 / 255, 128 / 255, 144 / 255)),
    ]

    random.shuffle(couleur)
    couleur = [
        ([0, 255, 255, 0], "#ffff00", (255 / 255, 255 / 255, 0)),
        ([0, 215, 255, 0], "#ffd700", (255 / 255, 215 / 255, 0)),
    ] + couleur

    for i in initiateurs:
        co = random.randrange(0, 255)
        dicocouleur[i] = couleur[listechaine][0]
        dicomoyennefabric1[i] = []
        dicomoyennefabric2[i] = []
        dicochaine[i] = []
        listechaine += 1
    listechaine = 0
    while len(chainefini) <= 200 and tempo < tstochastic:
        r = []
        rlabel = []
        for i in initiateurs:
            if i == "I-I":
                r.append(initiateurs[i][0] * initiateurs[i][1])
                rlabel.append("d")
                if (initiateurs[i][0] * initiateurs[i][1]) < 0:
                    exit()
            elif i == "I°":
                r.append(
                    initiateurs[i][0] * initiateurs[i][1] * (initiateurs[i][1] - 1)
                )
                rlabel.append("iI°")
            else:
                r.append(initiateurs[i][0] * initiateurs[i][1] * initiateurs["I°"][1])
                rlabel.append("i" + i)

        for i in reactants:
            r.append(i[1] * len(dicochaine[i[0][0]]) * initiateurs[i[0][1]][1])
            rlabel.append("p" + i[0][0] + "+" + i[0][1])

        r.append(terminaison[0] * listechaine * (listechaine - 1))
        rlabel.append("tf")  # fusion terminaison
        r.append(terminaison[1] * listechaine * (listechaine - 1))
        rlabel.append("tp")  # partage terminaison

        sommeratio = sum(r)

        unitint = [0]
        for i in r:
            unitint.append(unitint[-1] + i / sommeratio * 100)
        unitint.pop(0)

        choix = ""
        s1 = random.randrange(0, int(unitint[-1]))
        for i in range(len(unitint)):
            if unitint[i] > s1:
                choix = rlabel[i]
                break

        if choix == "d":
            initiateurs["I°"] = (initiateurs["I°"][0], initiateurs["I°"][1] + 2)
            initiateurs["I-I"] = (initiateurs["I-I"][0], initiateurs["I-I"][1] - 1)
        elif choix == "iI°":
            initiateurs["I°"] = (initiateurs["I°"][0], initiateurs["I°"][1] - 2)
        elif choix[0] == "i":
            initiateurs["I°"] = (initiateurs["I°"][0], initiateurs["I°"][1] - 1)
            initiateurs[choix[1:]] = (
                initiateurs[choix[1:]][0],
                initiateurs[choix[1:]][1] - 1,
            )
            dicochaine[choix[1:]].append("i_" + choix[1:])
            listechaine += 1
            dicomoyennefabric1[choix[1:]].append(1)
        elif choix[0] == "p":
            monomere = choix.split("+")[1]
            chaine = choix.split("+")[0][1:]
            initiateurs[monomere] = (
                initiateurs[monomere][0],
                initiateurs[monomere][1] - 1,
            )

            s2 = random.randrange(len(dicochaine[chaine]))
            chainechoisi = dicochaine[chaine].pop(s2)
            if chaine == monomere:
                dicomoyennefabric1[chaine].append(
                    dicomoyennefabric1[chaine].pop(s2) + 1
                )
            else:
                dicomoyennefabric2[chaine].append(dicomoyennefabric1[chaine].pop(s2))
                dicomoyennefabric1[monomere].append(1)
            dicochaine[monomere].append(chainechoisi + "_" + monomere)
        elif choix[0] == "t":
            s2 = random.randrange(listechaine)
            for i in dicochaine:
                if s2 >= len(dicochaine[i]):
                    s2 -= len(dicochaine[i])

                else:
                    chainechoisi1 = dicochaine[i].pop(s2)
                    monochoi1 = i
                    break

            listechainechois1 = chainechoisi1.split("_")
            chainechoisi1 = ""
            dicomoyennefabric2[monochoi1].append(dicomoyennefabric1[monochoi1].pop(s2))
            for ii in listechainechois1:
                chainechoisi1 += ii
            listechaine -= 1
            s2 = random.randrange(listechaine)
            for i in dicochaine:
                if s2 >= len(dicochaine[i]):
                    s2 -= len(dicochaine[i])
                else:
                    chainechoisi2 = dicochaine[i].pop(s2)
                    monochoi2 = i
                    break

            listechaine -= 1
            if choix[1] == "f":
                dicomoyennefabric2[monochoi1][-1] += dicomoyennefabric1[monochoi2].pop(
                    s2
                )
                listechainechois2 = chainechoisi2.split("_")
                listechainechois2.reverse()
                chainechoisi2 = ""
                for ii in listechainechois2:
                    chainechoisi2 += ii
                chainefini.append(chainechoisi1 + chainechoisi2)
                initiateur(f)
                for i in listechainechois1[1:]:
                    monomerec(f, dicocouleur[i])
                for i in listechainechois2[:-1]:
                    monomerec(f, dicocouleur[i])
                initiateur(f)
                fini_ligne(f, tranche)
                chainechoisif.append(listechainechois1 + listechainechois2)

            elif choix[1] == "p":
                listechainechois2 = chainechoisi2.split("_")
                dicomoyennefabric2[monochoi2].append(
                    dicomoyennefabric1[monochoi2].pop(s2)
                )
                chainechoisi2 = ""
                for ii in listechainechois2:
                    chainechoisi2 += ii
                chainefini.append(chainechoisi1)
                chainefini.append(chainechoisi2)
                initiateur(f)
                for i in listechainechois1[1:]:
                    monomerec(f, dicocouleur[i])
                fini_ligne(f, tranche)
                initiateur(f)
                for i in listechainechois2[1:]:
                    monomerec(f, dicocouleur[i])
                fini_ligne(f, tranche)
                chainechoisif.append(listechainechois1.copy())
                chainechoisif.append(listechainechois2.copy())

        tau = -log(random.random()) / sommeratio
        tempo += tau

    fini_image(f, tranche)
    commence_image(g)

    chainechoisif.sort(key=len)
    chainechoisif.reverse()
    for i in chainechoisif:
        for j in i:
            if j == "i":
                initiateur(g)
            else:
                monomerec(g, dicocouleur[j])
        fini_ligne(g, tranche)
    fini_image(g, tranche)
    f.close()
    g.close()
    t2 = t.time()
    retour = []
    for key in dicomoyennefabric2:
        if len(dicomoyennefabric2[key]) != 0:
            retour.append(
                (key, (sum(dicomoyennefabric2[key]) / len(dicomoyennefabric2[key])))
            )
    return (str(t2 - t1), retour, dicomoyennefabric2, couleur)


def simulanal(tranche: int, p11: float, p22: float):
    t1 = t.time()

    tha = 1 / (1 - p11)

    thb = 1 / (1 - p22)

    activemonomère = random.sample(["rouge", "bleu"], k=1)[0]
    listetailleb = []
    listetailler = []
    polymere = []
    for i in range(tranche * tranche):
        if activemonomère == "rouge":
            alea = random.randint(0, 10000)
            if alea < p11 * 10000:
                activemonomère = "rouge"
            else:
                activemonomère = "bleu"

        else:
            alea = random.randint(0, 10000)
            if alea < p22 * 10000:
                activemonomère = "bleu"
            else:
                activemonomère = "rouge"

        polymere.append(activemonomère)

    distance = 1
    for i in range(len(polymere) - 1):
        if polymere[i] == polymere[i + 1]:
            distance += 1
        else:
            if polymere[i] == "rouge":
                listetailler.append(distance)

            else:
                listetailleb.append(distance)
            distance = 1

    if polymere[i] == "rouge":
        listetailler.append(distance)

    if polymere[i] == "bleu":
        listetailleb.append(distance)

    listetailler.sort()

    listetailleb.sort()
    lenr = len(listetailler)
    resa = sum(listetailler) / lenr
    maxa = max(listetailler)
    maxb = max(listetailleb)
    mina = min(listetailler)
    minb = min(listetailleb)
    lenb = len(listetailleb)
    resb = sum(listetailleb) / lenb

    arepart = []
    for i in range(10):
        arepart.append(listetailler[int(lenr / 10 * i)])
    arepart.append(listetailler[-1])
    brepart = []
    for i in range(10):
        brepart.append(listetailleb[int(lenb / 10 * i)])
    brepart.append(listetailleb[-1])
    t2 = t.time()
    return (tha, resa, thb, resb, mina, maxa, minb, maxb, arepart, brepart, t2 - t1)


def fonction(co, A, x):
    fa0 = co
    rS = A[0]
    rPOT = A[1]
    alpha = rS / (1 - rS)
    beta = rPOT / (1 - rPOT)
    gamma = (1 - rS * rPOT) / ((1 - rS) * (1 - rPOT))
    delta = (1 - rS) / (2 - rS - rPOT)
    fa = x
    f = 1 - (fa / fa0) ** (alpha) * ((1 - fa) / (1 - fa0)) ** (beta) * (
        (fa0 - delta) / (fa - delta)
    ) ** (round(gamma, 6))
    return f


def moindre_carré(
    appli, con: list, xexp: list, fdotexpt: list, tranche1: range, tranche2: range
):
    ta = t.time()
    appli.temps = tk.Label(
        appli,
        text="temps de calcul: " + str(0) + "secondes",
        font="Helvetica 20 bold",
        bg="light green",
    )
    appli.char = tk.Label(
        appli, text="chargement", font="Helvetica 40 bold", bg="light green"
    )
    appli.char.place(x=300, y=150)
    appli.temps.place(x=250, y=300)
    appli.update_idletasks()
    t1 = t.time()
    liste = []
    for i in tranche1:
        for j in tranche2:
            if i == 100 or j == 100:
                pass
            elif i + j == 200:
                pass
            else:
                liste.append((i / 100, j / 100))
    distance = []
    fraction = len(liste) // 100
    fractionreste = len(liste) % 100
    for point in range(100):
        for k in liste[point * fraction : (point + 1) * fraction]:
            sommecarre = 0
            delta = (1 - k[0]) / (2 - k[0] - k[1])
            gamma = (1 - k[0] * k[1]) / ((1 - k[0]) * (1 - k[1]))
            for experience in range(len(fdotexpt)):
                for iter in range(len(fdotexpt[experience])):
                    if (
                        round(delta, 4) != round(fdotexpt[experience][iter], 4)
                        and round(delta, 4) != con[experience]
                    ):
                        if (con[experience] - delta) / (
                            fdotexpt[experience][iter] - delta
                        ) > 0:
                            carre = (
                                fonction(con[experience], k, fdotexpt[experience][iter])
                                - xexp[experience][iter]
                            )
                            if abs(carre) > 200000:
                                carre = 200000
                        else:
                            if round(gamma, 6) % 1 == 0:
                                carre = (
                                    fonction(
                                        con[experience], k, fdotexpt[experience][iter]
                                    )
                                    - xexp[experience][iter]
                                )
                            if abs(carre) > 200000:
                                carre = 200000
                            else:
                                carre = 200000
                    else:
                        carre = 200000

                    if isinstance(carre, complex):
                        carre = 200000
                    carre = carre**2
                    sommecarre += carre
            distance.append(sommecarre)
        tb = t.time()
        appli.temps.configure(
            text="temps de calcul: " + str(round(tb - ta, 1)) + "secondes"
        )
        appli.chargement.step(1)
        appli.update_idletasks()
    for k in liste[100 * fraction : 100 * fraction + fractionreste]:
        sommecarre = 0
        delta = (1 - k[0]) / (2 - k[0] - k[1])
        gamma = (1 - k[0] * k[1]) / ((1 - k[0]) * (1 - k[1]))
        for experience in range(len(fdotexpt)):
            for iter in range(len(fdotexpt[experience])):
                if (
                    round(delta, 4) != round(fdotexpt[experience][iter], 4)
                    and round(delta, 4) != con[experience]
                ):
                    if (con[experience] - delta) / (
                        fdotexpt[experience][iter] - delta
                    ) > 0:
                        carre = (
                            fonction(con[experience], k, fdotexpt[experience][iter])
                            - xexp[experience][iter]
                        )
                        if abs(carre) > 200000:
                            carre = 200000
                    else:
                        if round(gamma, 6) % 1 == 0:
                            carre = (
                                fonction(con[experience], k, fdotexpt[experience][iter])
                                - xexp[experience][iter]
                            )
                        if abs(carre) > 200000:
                            carre = 200000
                        else:
                            carre = 200000
                else:
                    carre = 200000

                if isinstance(carre, complex):
                    carre = 200000
                carre = carre**2
                sommecarre += carre
        distance.append(sommecarre)

    g = min(distance)

    for i in range(len(distance)):
        if distance[i] == g:
            minimumfind = liste[i]
    appli.chargement.place_forget()
    appli.char.place_forget()
    appli.temps.place_forget()
    appli.update_idletasks()
    return (minimumfind, g)


def calcpfa(fa, fa0, alpha, beta, delta, gamma):
    return 1 - (fa / fa0) ** (alpha) * ((1 - fa) / (1 - fa0)) ** (beta) * (
        (fa0 - delta) / (fa - delta)
    ) ** (gamma)


def log_tick_formatter(val, pos=None):
    return f"$10^{{{int(val)}}}$"


def dizaine_formatter(val, pos=None):
    return str(int(val)) + "0"


def polygon_under_graph(x, y):
    """
    Construct the vertex list which defines the polygon filling the space under
    the (x, y) line graph. This assumes x is in ascending order.
    """
    return [(0, 0), (x[0], 0.0), *zip(x, y), (x[-1], 0.0)]


def distribanal(r1, r2, f10):
    listfl = []
    listfal = []

    f20 = 1 - f10

    for jx in range(10):
        fa0 = f10

        alpha = r1 / (1 - r1)
        beta = r2 / (1 - r2)
        gamma = (1 - r1 * r2) / ((1 - r1) * (1 - r2))
        delta = (1 - r1) / (2 - r1 - r2)
        fa = np.arange(1000) / 1000
        conversion = 1 - (fa / fa0) ** (alpha) * ((1 - fa) / (1 - fa0)) ** (beta) * (
            (fa0 - delta) / (fa - delta)
        ) ** (round(gamma, 6))
        fal = list(fa)
        convl = list(conversion)
        dmin = abs(jx / 10 - convl[0])
        mini = 0
        for i in range(len(convl)):
            dist = abs(jx / 10 - convl[i])
            if dist < dmin:
                dmin = dist
                mini = i
        listfl.append(convl[mini])
        listfal.append(fal[mini])
    x = np.arange(0, 1000)
    y = []

    for i in listfal:
        m1 = 9500
        m2 = 500
        k11 = 1
        k12 = 2
        k22 = 2
        k21 = 1
        r1 = k11 / k12
        r2 = k22 / k21

        f1 = i
        f2 = 1 - f10

        p11 = f1 / (f1 + f2 / r1)

        p22 = f2 / (f2 + f1 / r2)
        z = p11 ** (x - 1) * (1 - p11) * x
        y.append(z)
    for i in listfal:
        m1 = 9500
        m2 = 500
        k11 = 1
        k12 = 1
        k22 = 2
        k21 = 1
        r1 = k11 / k12
        r2 = k22 / k21

        f1 = i
        f2 = 1 - f10

        p11 = f1 / (f1 + f2 / r1)

        p22 = f2 / (f2 + f1 / r2)
        z2 = p22 ** (x - 1) * (1 - p22) * x
        y.append(z2)
    return x, y


class experience:
    def __init__(self, appli, numéro, comps, comppot):
        self.comps = comps
        self.comppot = comppot
        self.rapportcanv2 = tk.Canvas(
            appli,
            bg="light green",
            height=110,
            width=800,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )
        self.appli = appli

        self.exptable1 = []
        self.exptable3 = []
        self.exptable2 = []
        self.expabs = []
        self.expord = []
        self.retour2contour = self.rapportcanv2.create_oval(
            10, 10, 70, 70, fill="black"
        )
        self.retour2 = self.rapportcanv2.create_line(
            20,
            40,
            60,
            40,
            arrow="first",
            arrowshape=(20, 20, 20),
            width=5,
            fill="light green",
            activefill="white",
        )

        self.rapportcanv2.bind("<ButtonRelease-1>  ", self.clickdroit)
        self.experiencetitre = self.rapportcanv2.create_text(
            400, 40, text="experience " + str(numéro), font=("times", "28", "bold")
        )
        self.mesuretext = tk.Label(
            appli,
            text="nombre de\n mesures:",
            font="Helvetica 18",
            bg="light green",
        )
        self.mesurecurseur = tk.Scale(
            appli,
            orient=tk.HORIZONTAL,
            from_=1,
            to=10,
            bg="light green",
            highlightthickness=0,
            troughcolor="white",
            activebackground="light grey",
            font="Helvetica 18",
            length=150,
            width=20,
        )
        self.mesurecurseur.set(1)
        self.mesureconfirm = tk.Button(
            appli,
            height=1,
            width=10,
            text="confirmer",
            font=("Helvetica 18 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.confirme,
        )

        self.rapportconfirme2 = tk.Button(
            appli,
            height=1,
            width=20,
            text="confirmer",
            font=("Helvetica 26 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.expsuivante,
        )
        self.temprec = self.rapportcanv2.create_rectangle(
            240, 109, 419, 83, fill="light blue", outline="grey"
        )
        self.temptext = self.rapportcanv2.create_text(
            332, 96, text="temps", font="Helvetica 14 "
        )
        self.comp1rec = self.rapportcanv2.create_rectangle(
            419, 109, 598, 83, fill="light blue", outline="grey"
        )
        self.comp1text = self.rapportcanv2.create_text(
            516, 96, text="conversion " + self.comps, font="Helvetica 14 "
        )
        self.comp2rec = self.rapportcanv2.create_rectangle(
            598, 109, 777, 83, fill="light blue", outline="grey"
        )
        self.comp1text = self.rapportcanv2.create_text(
            700, 96, text="conversion " + self.comppot, font="Helvetica 14 "
        )

        self.concintxt = tk.Label(
            appli,
            text="concentration\ninitiale de " + self.comppot + ":",
            font="Helvetica 18 ",
            bg="light green",
        )
        self.concintab = tk.Entry(self.appli, width=15, font="Helvetica 18 ")

    def clickdroit(self, mclick):
        self.appli.clickdroit(mclick)

    def expsuivante(self):
        try:
            self.concin = float(self.concintab.get())
            for i in range(len(self.exptable1)):
                convs = float(self.exptable3[i].get())
                convpot = float(self.exptable2[i].get())
                self.expabs.append(convs * (1 - self.concin) + convpot * (self.concin))
                # x expt
                self.expord.append(
                    (1 - convpot)
                    * self.concin
                    / ((1 - convpot) * self.concin + (1 - convs) * (1 - self.concin))
                )
                # fpott expt
            self.appli.expsuivante()
        except ValueError:
            Message(
                [
                    "erreur de synthaxe",
                    "sources d'erreurs possibles:",
                    "-vous avez peut-être oublié de remplir une case.",
                    "-vous avez peut-être confondu un point et une virgule:",
                    "4,2 -> synthaxe correcte: 4.2 (ne tapez jamais de virgules !!)",
                    "-vous avez peut-être mal écrit une forme exponentielle:",
                    "5.3*10^-5 -> synthaxe correcte: 5.3e-5",
                ],
                (self.appli.winfo_x(), self.appli.winfo_y()),
            )
        except ZeroDivisionError:
            Message(
                [
                    "erreur de calcul",
                    "sources d'erreurs possibles:",
                    "- le calcul avec ces données entraine une division par 0",
                ],
                (self.appli.winfo_x(), self.appli.winfo_y()),
            )

    def confirme(self):
        self.valeur = self.mesurecurseur.get()
        if self.valeur > len(self.exptable2):
            for i in range(self.valeur - len(self.exptable2)):
                self.exptable1.append(
                    tk.Entry(self.appli, width=16, font="Helvetica 15")
                )
                self.exptable3.append(
                    tk.Entry(self.appli, width=16, font="Helvetica 15")
                )
                self.exptable2.append(
                    tk.Entry(self.appli, width=16, font="Helvetica 15")
                )
        elif self.valeur < len(self.exptable2):
            for i in range(len(self.exptable2) - self.valeur):
                self.exptable1[-1].place_forget()

                self.exptable3[-1].place_forget()
                self.exptable2[-1].place_forget()
                del self.exptable1[-1]
                del self.exptable3[-1]
                del self.exptable2[-1]

        for i in range(len(self.exptable2)):
            self.exptable1[i].place(x=240, y=110 + 27 * i)
            self.exptable3[i].place(x=419, y=110 + 27 * i)
            self.exptable2[i].place(x=598, y=110 + 27 * i)

    def pack(self):
        self.mesurecurseur.place(x=10, y=150)
        self.mesureconfirm.place(x=10, y=210)
        self.mesuretext.place(x=10, y=95)
        self.rapportcanv2.pack()
        self.rapportconfirme2.place(x=200, y=400)
        self.concintxt.place(x=5, y=270)
        self.concintab.place(x=10, y=330)
        self.confirme()

    def pack_forget(self):
        self.mesurecurseur.place_forget()
        self.mesureconfirm.place_forget()
        self.mesuretext.place_forget()
        self.rapportcanv2.pack_forget()
        self.rapportconfirme2.place_forget()
        self.concintab.place_forget()
        self.concintxt.place_forget()
        for i in range(len(self.exptable2)):
            self.exptable1[i].place_forget()
            self.exptable2[i].place_forget()
            self.exptable3[i].place_forget()


class experience2:
    def __init__(self, fichier, rangé) -> None:
        sheet_obj = fichier.active
        cellconcin = sheet_obj.cell(row=rangé, column=12)

        self.concin = float(cellconcin.value)

        self.expabs = [0]
        self.expord = [self.concin]

        for i in range(8):
            cellconvs = sheet_obj.cell(row=rangé + i, column=13)
            cellconvpot = sheet_obj.cell(row=rangé + i, column=14)

            convs = float(cellconvs.value)
            convpot = float(cellconvpot.value)
            if convs != 0.0 and convpot != 0.0:
                self.expabs.append(convs * (1 - self.concin) + convpot * (self.concin))
                # x expt
                self.expord.append(
                    (1 - convpot)
                    * self.concin
                    / ((1 - convpot) * self.concin + (1 - convs) * (1 - self.concin))
                )
                # fpott expt


class Message(tk.Tk):
    def __init__(self, message, coordinate):
        tk.Tk.__init__(
            self,
        )
        self.enrv = 0
        if isinstance(message, str):
            texte = message
            soustexte = [[]]
        elif isinstance(message, list):
            texte = message[0]
            soustexte = message[1:]
        self.geometry(
            str(max(len(texte) * 20, len(max(soustexte, key=len)) * 8))
            + "x"
            + str(90 + len(soustexte) * 16)
            + "+"
            + str(coordinate[0] + 200)
            + "+"
            + str(coordinate[1] + 250)
        )
        self.wm_overrideredirect(True)
        w = tk.Frame(
            self,
            bg="khaki1",
            height=90 + len(soustexte) * 16,
            width=max(len(texte) * 20, len(max(soustexte, key=len)) * 8),
            highlightcolor="black",
            highlightthickness=4,
            highlightbackground="black",
        )

        self.label = tk.Label(
            self,
            bg="khaki1",
            text=texte,
            font=("Helvetica 17 bold"),
            foreground="dark blue",
        )
        self.label2 = []
        for i in soustexte:
            self.label2.append(
                tk.Label(self, bg="khaki1", text=i, font=("Helvetica 10 "))
            )
        j = 0
        for i in self.label2:
            i.place(relx=0.5, y=45 + 16 * j, anchor=tk.CENTER)
            j += 1
        self.bout = tk.Button(
            self,
            text="OK",
            font=("Helvetica 16 bold"),
            command=self.destroy,
            background="light blue",
        )
        w.place(x=0, y=0)
        self.bind("<Button-1>", self.select)
        self.bind("<B1-Motion>", self.move)
        self.label.place(relx=0.5, y=20, anchor=tk.CENTER)

        self.bout.place(relx=0.5, y=60 + 16 * len(soustexte), anchor=tk.CENTER)
        self.resizable(False, False)
        self.configure(bg="khaki1")
        self.title("alerte")
        self.mainloop()

    def select(self, mclick):
        self.mclick = (
            self.winfo_pointerx() - self.winfo_rootx(),
            self.winfo_pointery() - self.winfo_rooty(),
        )

    def move(self, mclick):
        x, y = (
            self.winfo_pointerx() - self.mclick[0],
            self.winfo_pointery() - self.mclick[1],
        )

        self.geometry(f"+{x}+{y}")


def Question(parent, texte, coordinate):
    my_w_child = tk.Toplevel(parent)  # Child window
    my_w_child.geometry(
        str(8 + len(texte) * 20)
        + "x138+"
        + str(coordinate[0] + 200)
        + "+"
        + str(coordinate[1] + 250)
    )  # Size of the window
    my_w_child.resizable(False, False)
    w = tk.Frame(
        my_w_child,
        bg="khaki1",
        height=138,
        width=8 + len(texte) * 20,
        highlightcolor="black",
        highlightthickness=4,
        highlightbackground="black",
    )
    w.pack()
    l1 = tk.Label(
        my_w_child,
        bg="khaki1",
        text=texte,
        font="Helvetica 17 bold",
        foreground="dark blue",
    )
    l1.place(relx=0.5, y=20, anchor=tk.CENTER)
    my_w_child.wm_overrideredirect(True)
    my_w_child.configure(bg="khaki1")
    e1 = tk.Entry(my_w_child, width=len(texte) - 5, font="Helvetica 16 ")
    e1.insert(tk.END, "nom")
    e1.place(relx=0.5, y=57, anchor=tk.CENTER)

    b2 = tk.Button(
        my_w_child,
        text="confirmer",
        font=("Helvetica 17 bold"),
        command=lambda: [parent.addmonof2(e1.get()), my_w_child.destroy()],
        background="light blue",
    )
    b2.place(relx=0.5, y=100, anchor=tk.CENTER)

    def select(mclick):
        my_w_child.mclick = (
            my_w_child.winfo_pointerx() - my_w_child.winfo_rootx(),
            my_w_child.winfo_pointery() - my_w_child.winfo_rooty(),
        )

    def move(mclick):
        if (
            my_w_child.mclick[0] < 115
            or my_w_child.mclick[0] > 445
            or my_w_child.mclick[1] < 45
            or my_w_child.mclick[1] > 75
        ):
            x, y = (
                my_w_child.winfo_pointerx() - my_w_child.mclick[0],
                my_w_child.winfo_pointery() - my_w_child.mclick[1],
            )
            my_w_child.geometry(f"+{x}+{y}")

    my_w_child.bind(
        "<Return>", lambda x: [parent.addmonof2(e1.get()), my_w_child.destroy()]
    )
    my_w_child.bind("<Button-1>", select)
    my_w_child.bind("<B1-Motion>", move)


#########################################################################################
#   _______  _______  ___      __   __  _______  __   __  _______  __   __  _______     #
#   |       ||       ||   |    |  | |  ||       ||  | |  ||       ||  |_|  ||       |   #
#   |    _  ||   _   ||   |    |  |_|  ||    _  ||  |_|  ||    ___||       ||    ___|   #
#   |   |_| ||  | |  ||   |    |       ||   |_| ||       ||   |___ |       ||   |___    #
#   |    ___||  |_|  ||   |___ |_     _||    ___||       ||    ___||       ||    ___|   #
#   |   |    |       ||       |  |   |  |   |    |   _   ||   |___ | ||_|| ||   |___    #
#   |___|    |_______||_______|  |___|  |___|    |__| |__||_______||_|   |_||_______|   #
#                                                                                       #
#########################################################################################
#    .----------------.  .----------------.  .----------------.     #
#   | .--------------. || .--------------. || .--------------. |    #
#   | | ____   ____  | || |     ____     | || |  ____  ____  | |    #
#   | ||_  _| |_  _| | || |   .'    `.   | || | |_  _||_  _| | |    #
#   | |  \ \   / /   | || |  /  .--.  \  | || |   \ \  / /   | |    #
#   | |   \ \ / /    | || |  | |    | |  | || |    > `' <    | |    #
#   | |    \ ' /     | || |  \  `--'  /  | || |  _/ /'`\ \_  | |    #
#   | |     \_/      | || |   `.____.'   | || | |____||____| | |    #
#   | |              | || |              | || |              | |    #
#   | '--------------' || '--------------' || '--------------' |    #
#    '----------------'  '----------------'  '----------------'     #
#                                                                   #
#####################################################################


class aplli(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        ######################################################################
        #                paramétrage de la fenetre                           #
        ######################################################################
        self.geometry("800x500")
        self.resizable(False, False)
        self.configure(bg="light green")
        self.style = ttk.Style(self)

        self.style.theme_create(
            "ussop",
            settings={
                "TMenubutton": {"configure": {"background": "light pink"}},
                "TButton": {"configure": {"background": "light pink"}},
                "TFrame": {"configure": {"background": "light green"}},
                "TLabel": {"configure": {"background": "light green"}},
                "TScrollbar": {"configure": {"background": "light pink"}},
                "TCombobox": {"configure": {"background": "light blue"}},
                "TProgressbar": {
                    "configure": {
                        "background": "light blue",
                        "darkcolor": "light grey",
                        "bordercolor": "light grey",
                        "foreground": "light grey",
                        "thickness": "50",
                    }
                },
            },
        )

        self.style.theme_use("ussop")
        ######################################################################
        #                             menu                                   #
        ######################################################################
        self.canvmenu = tk.Canvas(
            self,
            bg="light green",
            height=230,
            width=800,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )
        self.frame = tk.Frame(self, width=200, height=200)
        self.frame.pack_propagate(0)
        img = Image.open("logo.png")
        self.img = ImageTk.PhotoImage(img)
        self.label = tk.Label(self.frame, image=self.img, background="light green")
        self.titre = self.canvmenu.create_text(
            500, 100, text="Polyphème", font=("times", "55", "bold")
        )
        self.rapportbout = tk.Button(
            self,
            height=1,
            width=20,
            text="rapports réactionnels",
            font=("Helvetica 20 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.rapport1,
        )
        self.analbout = tk.Button(
            self,
            height=1,
            width=23,
            text="simulation analytique",
            font=("Helvetica 20 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.simul1,
        )
        self.mcbout = tk.Button(
            self,
            height=1,
            width=23,
            text="simulation Monte-Carlo",
            font=("Helvetica 20 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.mc1,
        )
        self.canvmenu.pack()
        self.rapportbout.pack()
        self.analbout.place(x=400, y=325, anchor=tk.CENTER)
        self.mcbout.place(x=400, y=395, anchor="center")
        self.frame.place(x=100, y=10)
        self.label.pack()
        ######################################################################
        #              rapport de réaction -entrées                          #
        ######################################################################
        self.composéAcell = tk.Entry(self, width=4, font="Helvetica 15")
        self.composéBcell = tk.Entry(self, width=4, font="Helvetica 15")
        self.composélabel = tk.Label(
            self, text="noms des composés:", font="Helvetica 18 bold", bg="light green"
        )
        self.composéAcell.insert(tk.END, "A")
        self.composéBcell.insert(tk.END, "B")
        self.rapportcanv1 = tk.Canvas(
            self,
            bg="light green",
            height=80,
            width=80,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )
        self.retour1contour = self.rapportcanv1.create_oval(
            10, 10, 70, 70, fill="black"
        )
        self.retour1 = self.rapportcanv1.create_line(
            20,
            40,
            60,
            40,
            arrow="first",
            arrowshape=(20, 20, 20),
            width=5,
            fill="light green",
            activefill="white",
        )
        self.rapportcanv1.bind("<ButtonRelease-1>  ", self.clickdroit)
        self.rapportcanv12 = tk.Canvas(
            self,
            bg="light green",
            height=420,
            width=800,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )
        self.rapportcanv12.create_line(0, 100, 800, 100, width=3)
        self.rapportcanv12.create_line(0, 3, 800, 3, width=3)
        self.rapportcanv12.create_line(400, 3, 400, 420, width=3)
        self.manuel = tk.Label(
            self,
            text="mode manuel:",
            font="Helvetica 25 bold",
            bg="light green",
        )
        self.experiencetexte = tk.Label(
            self,
            text="nombre \nd'expériences:",
            font="Helvetica 25",
            bg="light green",
        )
        self.importation = tk.Label(
            self,
            text="importer \ndepuis excel:",
            font="Helvetica 25 bold",
            bg="light green",
        )
        self.experiencecurseur = tk.Scale(
            self,
            orient=tk.HORIZONTAL,
            from_=1,
            to=10,
            bg="light green",
            highlightthickness=0,
            troughcolor="white",
            activebackground="light grey",
            font="Helvetica 25",
            length=150,
            width=20,
        )
        self.rapportconfirme1 = tk.Button(
            self,
            height=1,
            width=10,
            text="confirmer",
            font=("Helvetica 26 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.rapport2,
        )
        self.rapportconfirme2 = tk.Button(
            self,
            height=1,
            width=10,
            text="confirmer",
            font=("Helvetica 26 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.rapportexcel,
        )
        self.frame3 = tk.Frame(self, width=85, height=82)
        self.frame3.pack_propagate(0)
        self.img3 = tk.PhotoImage(file="parametre.png", width=85, height=82)
        self.label3 = tk.Label(self.frame3, image=self.img3, background="light green")
        self.label3.bind("<ButtonRelease-1>", self.parametre)
        self.cheminextext = tk.Label(
            self,
            text="emplacement:",
            font="Helvetica 25 ",
            bg="light green",
        )
        self.cheminex = tk.Entry(self, width=24, font="Helvetica 20 ")
        self.cheminboutonex = tk.Button(
            self,
            height=1,
            width=8,
            text="chercher",
            command=self.browse_button2,
            font="Helvetica 20 ",
            background="light pink",
        )
        self.temps = tk.Label(
            self,
            text="temps de calcul: " + str(0) + "secondes",
            font="Helvetica 20 bold",
            bg="light green",
        )
        self.char = tk.Label(
            self, text="chargement", font="Helvetica 40 bold", bg="light green"
        )
        self.chargement = ttk.Progressbar(
            self,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            value=0,
            length=400,
        )

        ######################################################################
        #              rapport de réaction -parametre                        #
        ######################################################################

        self.tranchsliders = tk.Canvas(
            self,
            bg="light green",
            height=42,
            width=442,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )
        self.slidersligne = self.tranchsliders.create_line(20, 20, 420, 20, width=5)
        self.sliderslignep = self.tranchsliders.create_line(
            20, 20, 170, 20, fill="white", width=3
        )
        self.tranchcurseurs1 = self.tranchsliders.create_oval(
            0, 0, 40, 40, fill="light blue", outline="grey"
        )
        self.tranchcurseurs2 = self.tranchsliders.create_oval(
            150, 0, 190, 40, fill="light blue", outline="grey"
        )
        self.placecurs1 = 0
        self.placecurs2 = 150
        self.curseurselect = ""
        self.tranchsliders.bind("<Button-1>", self.checkscurseurt)
        self.tranchsliders.bind("<ButtonRelease-1>", self.freescurseurt)
        self.tranchsliders.bind("<B1-Motion>", self.updatescurseurt)

        self.tranchsliderpot = tk.Canvas(
            self,
            bg="light green",
            height=42,
            width=442,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )
        self.sliderpotligne = self.tranchsliderpot.create_line(20, 20, 420, 20, width=5)
        self.sliderpotlignep = self.tranchsliderpot.create_line(
            20, 20, 170, 20, fill="white", width=3
        )
        self.tranchcurseurpot1 = self.tranchsliderpot.create_oval(
            0, 0, 40, 40, fill="light blue", outline="grey"
        )
        self.tranchcurseurpot2 = self.tranchsliderpot.create_oval(
            150, 0, 190, 40, fill="light blue", outline="grey"
        )
        self.placecurpot1 = 0
        self.placecurpot2 = 150
        self.curseurselect = ""
        self.tranchsliderpot.bind("<Button-1>", self.checkpotcurseurt)
        self.tranchsliderpot.bind("<ButtonRelease-1>", self.freepotcurseurt)
        self.tranchsliderpot.bind("<B1-Motion>", self.updatepotcurseurt)

        self.tranchslabel = tk.Label(
            self,
            text="zone de recherche \ndu composé " + "A" + ":",
            font="Helvetica 20",
            bg="light green",
        )
        self.tranchpotlabel = tk.Label(
            self,
            text="zone de recherche \ndu composé " + "B" + ":",
            font="Helvetica 20",
            bg="light green",
        )
        self.tranches1lab = tk.Label(
            self,
            text="0",
            font="Helvetica 15",
            bg="light green",
        )
        self.debtranchs = 0
        self.tranches2lab = tk.Label(
            self,
            text="3",
            font="Helvetica 15 ",
            bg="light green",
        )
        self.fintranchs = 3
        self.tranchepot1lab = tk.Label(
            self,
            text="0",
            font="Helvetica 15 ",
            bg="light green",
        )
        self.debtranchpot = 0
        self.tranchepot2lab = tk.Label(
            self,
            text="3",
            font="Helvetica 15",
            bg="light green",
        )
        self.fintranchpot = 3

        self.precslider = tk.Canvas(
            self,
            bg="light green",
            height=42,
            width=280,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )
        self.sliderprecligne = self.precslider.create_line(19, 20, 200, 20, width=5)
        self.sliderpreclignep = self.precslider.create_line(
            20, 20, 200, 20, fill="white", width=3
        )
        self.tranchcurseurprec = self.precslider.create_oval(
            180, 0, 220, 40, fill="light blue", outline="grey"
        )
        self.placecurprec = 180

        self.precslider.create_line(250, 15, 280, 15, width=2)
        self.precslider.create_line(265, 0, 265, 30, width=2)

        self.precslider.create_line(250, 39, 280, 39, width=2)

        self.preclabval = 0.01
        self.preclab = tk.Label(
            self,
            text="0.01",
            font="Helvetica 40 bold",
            bg="light green",
        )
        self.precslider.bind("<Button-1>", self.checkpreccurseurt)
        self.precslider.bind("<ButtonRelease-1>", self.freepreccurseurt)
        self.precslider.bind("<B1-Motion>", self.updatepreccurseurt)

        self.preclabel = tk.Label(
            self,
            text="précision:",
            font="Helvetica 20",
            bg="light green",
        )
        self.estimation = tk.Label(
            self,
            text="estimation du temps de calculs: \n"
            + "0"
            + "min "
            + "10"
            + "s "
            + "00",
            font="Helvetica 25 bold",
            bg="light green",
        )
        ######################################################################
        #              rapport de réaction -experience                       #
        ######################################################################
        self.exp = [experience(self, 1, "A", "B")]
        ######################################################################
        #              rapport de réaction -affichage                        #
        ######################################################################

        self.enrv = 0

        self.menubouton = tk.Button(
            self,
            height=1,
            width=10,
            text="quitter",
            font=("Helvetica 22 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.fini,
        )
        self.frame2 = tk.Frame(self, width=200, height=200)
        self.frame2.pack_propagate(0)
        self.img2 = tk.PhotoImage(file="save2.png", width=200, height=200)
        self.label2 = tk.Label(self.frame2, image=self.img2, background="light green")
        self.label2.bind("<ButtonRelease-1>", self.enregistre)
        self.canv3 = tk.Canvas(
            self,
            bg="light green",
            height=500,
            width=800,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )
        ######################################################################
        #              rapport de réaction -enregistrement                   #
        ######################################################################
        self.menuderoul = ttk.Combobox(
            values=[
                "tableur (format xlsx)",
                "image (format png)",
                "image (format jpg)",
                "image (format pdf)",
            ],
            width=22,
            font="Helvetica 17 ",
        )
        self.format = tk.Label(
            self,
            text="type de doc:",
            font="Helvetica 20 ",
            bg="light green",
        )
        self.nom = tk.Label(
            self,
            text="nom:",
            font="Helvetica 20 ",
            bg="light green",
        )
        self.nomtabl = tk.Entry(self, width=20, font="Helvetica 17 ")
        self.chemintext = tk.Label(
            self,
            text="emplacement:",
            font="Helvetica 20 ",
            bg="light green",
        )
        self.chemin = tk.Entry(self, width=40, font="Helvetica 17 ")
        self.cheminbouton = tk.Button(
            self,
            text="chercher",
            command=self.browse_button,
            font="Helvetica 15 bold",
            background="light pink",
        )
        self.enregistrebouton = tk.Button(
            self,
            text="sauvegarder",
            command=self.save,
            font="Helvetica 20 bold",
            activebackground="#D6E8F2",
            background="#AECFEB",
        )
        ######################################################################
        #              simulation analytique -entrées                        #
        ######################################################################
        self.rapportcanv22 = tk.Canvas(
            self,
            bg="light green",
            height=420,
            width=800,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )
        self.rapportcanv22.create_line(0, 3, 800, 3, width=3)
        self.rapportcanv22.create_line(400, 3, 400, 420, width=3)

        self.simulconfirme11 = tk.Button(
            self,
            height=1,
            width=10,
            text="confirmer",
            font=("Helvetica 26 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.simul21,
        )
        self.simulconfirme12 = tk.Button(
            self,
            height=1,
            width=10,
            text="confirmer",
            font=("Helvetica 26 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.simul22,
        )

        self.k11cell = tk.Entry(self, width=5, font="Helvetica 20 ")
        self.k12cell = tk.Entry(self, width=5, font="Helvetica 20 ")
        self.k21cell = tk.Entry(self, width=5, font="Helvetica 20 ")
        self.k22cell = tk.Entry(self, width=5, font="Helvetica 20 ")
        self.M1cell = tk.Entry(self, width=5, font="Helvetica 20 ")
        self.M2cell = tk.Entry(self, width=5, font="Helvetica 20 ")
        self.r1cell = tk.Entry(self, width=5, font="Helvetica 20 ")
        self.r2cell = tk.Entry(self, width=5, font="Helvetica 20 ")
        self.f1cell = tk.Entry(self, width=5, font="Helvetica 20 ")

        self.f1lab = tk.Label(
            self,
            text="f1:",
            font="Helvetica 25 ",
            bg="light green",
        )
        self.r1lab = tk.Label(
            self,
            text="r1:",
            font="Helvetica 25 ",
            bg="light green",
        )
        self.r2lab = tk.Label(
            self,
            text="r2:",
            font="Helvetica 25 ",
            bg="light green",
        )
        self.M1lab = tk.Label(
            self,
            text="M1:",
            font="Helvetica 25 ",
            bg="light green",
        )
        self.M2lab = tk.Label(
            self,
            text="M2:",
            font="Helvetica 25 ",
            bg="light green",
        )
        self.k11lab = tk.Label(
            self,
            text="k11:",
            font="Helvetica 25 ",
            bg="light green",
        )
        self.k12lab = tk.Label(
            self,
            text="k12:",
            font="Helvetica 25 ",
            bg="light green",
        )
        self.k21lab = tk.Label(
            self,
            text="k21:",
            font="Helvetica 25 ",
            bg="light green",
        )
        self.k22lab = tk.Label(
            self,
            text="k22:",
            font="Helvetica 25 ",
            bg="light green",
        )
        ######################################################################
        #              simulation analytique -théorie                        #
        ######################################################################
        self.simullance = tk.Button(
            self,
            height=1,
            width=20,
            text="lancer une simulation",
            font=("Helvetica 20 "),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.simul3,
        )
        self.distriblance = tk.Button(
            self,
            height=1,
            width=12,
            text="distribution",
            font=("Helvetica 20 "),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.simuldist,
        )
        self.result = tk.Label(
            self,
            text="yannis on t'aime",
            font="Helvetica 25",
            bg="light green",
        )
        ######################################################################
        #              simulation analytique -simulation                     #
        ######################################################################

        self.resultsim = tk.Label(
            self,
            text="usopp est un gros blaireau",
            font="Helvetica 25",
            bg="light green",
        )
        self.reparta = tk.Canvas(
            self,
            bg="white",
            width=300,
            height=20,
            highlightcolor="black",
            highlightbackground="black",
            highlightthickness=2,
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.repartb = tk.Canvas(
            self,
            bg="white",
            width=300,
            height=20,
            highlightcolor="black",
            highlightbackground="black",
            highlightthickness=2,
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.repartlabcanv = tk.Canvas(
            self,
            bg="light green",
            width=800,
            height=80,
            highlightthickness=0,
            borderwidth=0,
            relief=tk.FLAT,
        )

        self.monolabel = tk.Label(
            self, font=("Helvetica 16"), text="monomère", bg="light green"
        )
        self.frameanal1 = tk.Frame(self, width=200, height=200)
        self.frameanal1.pack_propagate(0)
        self.imganal1 = tk.PhotoImage(file="save2.png", width=200, height=200)
        self.labelanal1 = tk.Label(
            self.frameanal1, image=self.imganal1, background="light green"
        )
        self.labelanal1.bind("<ButtonRelease-1>", self.enregistre)

        ######################################################################
        #                      simulation MC -entrées                        #
        ######################################################################

        self.addmono = tk.Canvas(
            self,
            bg="light green",
            width=80,
            height=80,
            highlightthickness=0,
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.addcontour = self.addmono.create_oval(
            20, 20, 60, 60, activefill="white", width=3
        )
        self.add = self.addmono.create_line(
            40,
            25,
            40,
            55,
            width=3,
            fill="black",
        )
        self.add = self.addmono.create_line(
            25,
            40,
            55,
            40,
            width=3,
            fill="black",
        )
        self.addmono.bind("<Button-1>", self.addmonof)
        self.erasemono = tk.Canvas(
            self,
            bg="light green",
            width=80,
            height=80,
            highlightthickness=0,
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.erasecontour = self.erasemono.create_oval(
            20, 20, 60, 60, activefill="white", width=3
        )
        self.erase = self.erasemono.create_line(
            25,
            40,
            55,
            40,
            width=3,
            fill="black",
        )
        self.erasemono.bind("<Button-1>", self.deletemono)
        self.mcconfirm = tk.Button(
            self,
            height=2,
            width=12,
            text="lancer la\nsimulation",
            font=("Helvetica 20 bold"),
            activebackground="#D6E8F2",
            background="#AECFEB",
            command=self.simulationmc,
        )

        self.monolist = []
        ######################################################################
        #                          k propagation                             #
        ######################################################################
        self.monoprop = tk.Canvas(
            self,
            bg="DarkSeaGreen1",
            width=410,
            height=140,
            highlightthickness=1,
            highlightbackground="darkblue",
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.monoprop.create_text(
            200,
            10,
            text="constantes de vitesse de propagation:",
            font=("Helvetica 15 bold"),
        )
        self.monoprop.create_line(0, 20, 410, 20, fill="darkblue")
        x1 = 400
        x2 = 405
        y1 = 25
        y2 = 135
        r = 4
        points = (
            x1 + r,
            y1,
            x1 + r,
            y1,
            x2 - r,
            y1,
            x2 - r,
            y1,
            x2,
            y1,
            x2,
            y1 + r,
            x2,
            y1 + r,
            x2,
            y2 - r,
            x2,
            y2 - r,
            x2,
            y2,
            x2 - r,
            y2,
            x2 - r,
            y2,
            x1 + r,
            y2,
            x1 + r,
            y2,
            x1,
            y2,
            x1,
            y2 - r,
            x1,
            y2 - r,
            x1,
            y1 + r,
            x1,
            y1 + r,
            x1,
            y1,
        )

        self.monoprop.bind("<B1-Motion>", lambda event: self.curseurmc("1", event))
        self.polyprop = self.monoprop.create_polygon(
            points, smooth=True, fill="sea green"
        )
        self.monoproplistreac = []

        self.monoproplistreacell = []

        ######################################################################
        #                          k initialisation                          #
        ######################################################################
        self.monoinit = tk.Canvas(
            self,
            bg="DarkSeaGreen1",
            width=410,
            height=140,
            highlightthickness=1,
            highlightbackground="darkblue",
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.monoinit.create_text(
            200,
            10,
            text="constantes de vitesse des initiateurs:",
            font=("Helvetica 15 bold"),
        )
        self.monoinit.create_line(0, 20, 410, 20, fill="darkblue")
        x1 = 400
        x2 = 405
        y1 = 25
        y2 = 135
        r = 4
        points = (
            x1 + r,
            y1,
            x1 + r,
            y1,
            x2 - r,
            y1,
            x2 - r,
            y1,
            x2,
            y1,
            x2,
            y1 + r,
            x2,
            y1 + r,
            x2,
            y2 - r,
            x2,
            y2 - r,
            x2,
            y2,
            x2 - r,
            y2,
            x2 - r,
            y2,
            x1 + r,
            y2,
            x1 + r,
            y2,
            x1,
            y2,
            x1,
            y2 - r,
            x1,
            y2 - r,
            x1,
            y1 + r,
            x1,
            y1 + r,
            x1,
            y1,
        )

        self.monoinit.bind("<B1-Motion>", lambda event: self.curseurmc("0", event))
        self.polyinit = self.monoinit.create_polygon(
            points, smooth=True, fill="sea green"
        )
        self.monoinitlistreac = []
        self.monoinitlistreacell = []
        self.monoinitlistreac.append(
            tk.Label(
                self,
                font=("Helvetica 16"),
                text="I-I --> 2 I°",
                bg="DarkSeaGreen1",
            )
        )
        self.monoinitlistreac.append(
            tk.Label(
                self,
                font=("Helvetica 16 "),
                text="I°+I°--> Inerte",
                bg="DarkSeaGreen1",
            )
        )

        self.monoinitlistreacell.append(
            tk.Entry(self, width=7, font=("Helvetica 14"), background="DarkSeaGreen2")
        )
        self.monoinitlistreacell.append(
            tk.Entry(self, width=7, font=("Helvetica 14"), background="DarkSeaGreen2")
        )

        ######################################################################
        #                          k termination                             #
        ######################################################################
        self.monotermination = tk.Canvas(
            self,
            bg="DarkSeaGreen1",
            width=410,
            height=95,
            highlightthickness=1,
            highlightbackground="darkblue",
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.monotermination.create_text(
            200,
            10,
            text="constantes de vitesse de terminaison:",
            font=("Helvetica 15 bold"),
        )
        self.monotermination.create_line(0, 20, 410, 20, fill="darkblue")
        self.monotermlistreac = []

        self.monotermlistreac.append(
            tk.Label(
                self,
                font=("Helvetica 15"),
                text="Px°+Py°--> Px+y",
                bg="DarkSeaGreen1",
            )
        )
        self.monotermlistreac.append(
            tk.Label(
                self,
                font=("Helvetica 15"),
                text="Px°+Py°--> Px + Py",
                bg="DarkSeaGreen1",
            )
        )
        self.monotermlistreacell = []
        self.monotermlistreacell.append(
            tk.Entry(self, width=7, font=("Helvetica 14"), background="DarkSeaGreen2")
        )
        self.monotermlistreacell.append(
            tk.Entry(self, width=7, font=("Helvetica 14"), background="DarkSeaGreen2")
        )

        ######################################################################
        #                          concentration                             #
        ######################################################################

        self.monocon = tk.Canvas(
            self,
            bg="DarkSeaGreen1",
            width=345,
            height=170,
            highlightthickness=1,
            highlightbackground="darkblue",
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.monocon.create_text(
            170, 10, text="quantités de molécules:", font=("Helvetica 15 bold")
        )
        self.monocon.create_line(0, 20, 410, 20, fill="darkblue")
        x1 = 335
        x2 = 340
        y1 = 25
        y2 = 165
        r = 4
        points = (
            x1 + r,
            y1,
            x1 + r,
            y1,
            x2 - r,
            y1,
            x2 - r,
            y1,
            x2,
            y1,
            x2,
            y1 + r,
            x2,
            y1 + r,
            x2,
            y2 - r,
            x2,
            y2 - r,
            x2,
            y2,
            x2 - r,
            y2,
            x2 - r,
            y2,
            x1 + r,
            y2,
            x1 + r,
            y2,
            x1,
            y2,
            x1,
            y2 - r,
            x1,
            y2 - r,
            x1,
            y1 + r,
            x1,
            y1 + r,
            x1,
            y1,
        )
        self.monocon.bind("<B1-Motion>", lambda event: self.curseurmc("2", event))
        self.polyconc = self.monocon.create_polygon(
            points, smooth=True, fill="sea green"
        )

        self.monoconclistreac = []
        self.monoconclistreac.append(
            tk.Label(self, font=("Helvetica 15 "), text="I-I :", bg="DarkSeaGreen1")
        )

        self.monoconclistreac.append(
            tk.Label(self, font=("Helvetica 15 "), text="I°:", bg="DarkSeaGreen1")
        )

        self.monoconclistreacell = []
        self.monoconclistreacell.append(
            tk.Entry(self, width=7, font=("Helvetica 14 "), background="DarkSeaGreen2")
        )
        self.monoconclistreacell.append(
            tk.Entry(self, width=7, font=("Helvetica 14 "), background="DarkSeaGreen2")
        )

        ######################################################################
        #                          temps                                     #
        ######################################################################

        self.monotime = tk.Canvas(
            self,
            bg="DarkSeaGreen1",
            width=345,
            height=50,
            highlightthickness=1,
            highlightbackground="darkblue",
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.monotime.create_text(
            170, 10, text="temps stochastique:", font=("Helvetica 15 bold")
        )
        self.monotime.create_line(0, 20, 410, 20, fill="darkblue")
        self.monotimereac = self.monotime.create_text(
            10, 35, font=("Helvetica 15 "), text="\u03C4=", anchor=tk.W
        )
        self.monotimereacell = tk.Entry(
            self, width=6, font=("Helvetica 14 "), background="DarkSeaGreen2"
        )
        ######################################################################
        self.monoproplistfabric = []
        self.monomere = []
        self.inititer = 0
        self.propiter = 0
        self.conciter = 0

        ######################################################################
        #                          resultat mc                               #
        ######################################################################

        self.monoremc = tk.Canvas(
            self,
            bg="DarkSeaGreen1",
            width=445,
            height=260,
            highlightthickness=1,
            highlightbackground="darkblue",
            borderwidth=0,
            relief=tk.FLAT,
        )
        self.monoremc.create_text(
            220, 20, text="longueurs des chaines:", font=("Helvetica 25 bold")
        )
        self.monoremc.create_line(0, 40, 445, 40, fill="darkblue")
        self.monoremc.bind("<B1-Motion>", lambda event: self.curseurmc("2", event))
        self.variable2 = tk.StringVar(self, value="l")
        self.monolistlabel = []
        self.distribmc = tk.Radiobutton(
            self,
            variable=self.variable2,
            value="d",
            text="distribution",
            font="Helvetica 15 bold",
            activebackground="DarkSeaGreen2",
            background="DarkSeaGreen1",
            highlightbackground="darkblue",
            command=self.distribmclance,
        )
        self.radiomc = tk.Radiobutton(
            self,
            variable=self.variable2,
            value="l",
            text="longueurs des chaines",
            font="Helvetica 15 bold",
            activebackground="DarkSeaGreen2",
            background="DarkSeaGreen1",
            highlightbackground="darkblue",
            command=self.distriblanceoublie,
        )
        self.tranchechainne = 200
        self.limchaine = tk.Scale(
            self,
            label="limites de chaines produites",
            from_=100,
            to=1500,
            resolution=50,
            foreground="black",
            orient="horizontal",
            bg="DarkSeaGreen1",
            highlightthickness=1,
            highlightbackground="darkblue",
            troughcolor="white",
            activebackground="light grey",
            font="Helvetica 15 bold",
            length=300,
            width=20,
        )
        self.limchaine.set(200)
        self.framemc = tk.Frame(self, width=100, height=100)
        self.framemc.pack_propagate(0)
        self.imgmc = tk.PhotoImage(file="save2.png", width=100, height=100)
        self.labelmc = tk.Label(
            self.framemc, image=self.imgmc, background="light green"
        )
        self.labelmc.bind("<ButtonRelease-1>", self.enregistre)

    def menu(self):
        self.situation = "menu"
        self.canvmenu.pack()
        self.rapportbout.pack()
        self.analbout.place(x=400, y=325, anchor=tk.CENTER)
        self.mcbout.place(x=400, y=395, anchor=tk.CENTER)
        self.rapportcanv1.place_forget()
        self.addmono.place_forget()
        self.rapportcanv22.place_forget()
        self.rapportcanv12.place_forget()
        self.experiencetexte.place_forget()
        self.experiencecurseur.place_forget()
        self.rapportconfirme1.place_forget()
        self.k11cell.place_forget()
        self.k12cell.place_forget()
        self.k21cell.place_forget()
        self.k22cell.place_forget()
        self.M1cell.place_forget()
        self.M2cell.place_forget()
        self.f1cell.place_forget()
        self.r1cell.place_forget()
        self.r2cell.place_forget()
        self.f1lab.place_forget()
        self.r1lab.place_forget()
        self.r2lab.place_forget()
        self.M1lab.place_forget()
        self.M2lab.place_forget()
        self.k11lab.place_forget()
        self.k12lab.place_forget()
        self.k21lab.place_forget()
        self.k22lab.place_forget()
        self.importation.place_forget()
        self.manuel.place_forget()
        self.rapportconfirme2.place_forget()
        self.simulconfirme11.place_forget()
        self.simulconfirme12.place_forget()
        self.cheminex.place_forget()
        self.cheminboutonex.place_forget()
        self.cheminextext.place_forget()
        self.composélabel.place_forget()
        self.composéAcell.place_forget()
        self.composéBcell.place_forget()
        self.monoprop.place_forget()
        self.monoinit.place_forget()
        self.monotermination.place_forget()
        self.erasemono.place_forget()
        self.monocon.place_forget()
        self.monolabel.place_forget()
        self.limchaine.place_forget()

        self.label3.place_forget()
        self.frame3.place_forget()
        self.frame.place(x=100, y=10)
        self.label.pack()
        self.monolabel.place()
        self.monotime.place_forget()
        self.monotimereacell.place_forget()
        self.mcconfirm.place_forget()
        for i in range(min(len(self.monoinitlistreacell), 4)):
            self.monoinitlistreacell[i + self.inititer].place_forget()
            self.monoinitlistreac[i + self.inititer].place_forget()
        for i in range(min(len(self.monotermlistreacell), 4)):
            self.monotermlistreacell[i].place_forget()
            self.monotermlistreac[i].place_forget()
        for i in range(min(len(self.monoproplistreacell), 4)):
            self.monoproplistreacell[i + self.propiter].place_forget()
            self.monoproplistreac[i + self.propiter].place_forget()
        for i in range(min(len(self.monoconclistreacell), 5)):
            self.monoconclistreacell[i + self.conciter].place_forget()
            self.monoconclistreac[i + self.conciter].place_forget()

    def clickdroit(self, mclick):
        if (40 - mclick.y) ** 2 + (40 - mclick.x) ** 2 < 900:
            if self.situation == "rapport1":
                self.menu()
            elif self.situation == "affichemc":
                self.simulmc.place_forget()
                self.simulmc2.place_forget()
                self.distribmc.place_forget()
                self.radiomc.place_forget()
                self.framemc.place_forget()
                self.labelmc.pack_forget()
                try:
                    self.widget4.place_forget()
                except AttributeError:
                    pass
                self.mc1()
            elif self.situation == "mc1":
                self.menu()
            elif self.situation == "simul1":
                self.menu()
            elif self.situation == "simulaffiche":
                self.simul1()
            elif self.situation == "parametre":
                self.rapport1()
            elif self.situation == "simuldistrib":
                self.widget2.place_forget()
                self.widget3.place_forget()
                self.radio1.place_forget()
                self.frameanal1.place_forget()
                self.labelanal1.pack_forget()
                self.radio2.place_forget()
                self.cursor.place_forget()
                self.simul1()
            elif self.situation == "rapport2":
                if self.expact == 0:
                    self.rapport1()
                else:
                    self.experienceprecedente()
            elif self.situation == "telecharge1":
                self.affiche()
            elif self.situation == "telechargeanal":
                self.simuldist()
            elif self.situation == "telechargemc":
                self.affichemc()
            elif self.situation == "excel":
                self.exp = [experience(self, 1, self.comps, self.comppot)]
                self.rapportcanv1.pack_forget()
                self.menubouton.place_forget()
                self.reponse.place_forget()
                self.sommecarrereponse.place_forget()
                self.widget.place_forget()
                self.frame2.place_forget()
                self.label2.pack_forget()
                self.rapport1()
                self.curseurerreur.place_forget()
                self.ligne.place_forget()
                self.curseurlabel.place_forget()
                self.canvsliderpot.place_forget()
                self.canvsliders.place_forget()
                self.curseurlabel2.place_forget()
                # commande

    #################################################################
    #                                                               #
    #                simulation Monte Carlo                         #
    #                                                               #
    #################################################################

    def mc1(self):
        self.monolabel.place(x=185, y=25)
        self.inititer = 0
        self.propiter = 0
        self.conciter = 0
        self.situation = "mc1"
        self.rapportbout.pack_forget()
        self.analbout.place_forget()
        self.addmono.place(x=100, y=0)
        self.limchaine.place(x=460, y=300)
        self.mcconfirm.place(x=610, y=440, anchor=tk.CENTER)
        self.erasemono.place(x=300, y=0)
        self.monocon.place(x=440, y=50)
        self.mcbout.place_forget()
        self.canvmenu.pack_forget()
        self.rapportcanv1.place(x=0, y=0)
        self.monoprop.place(x=15, y=235)
        self.monoinit.place(x=15, y=79)
        self.monotermination.place(x=15, y=390)
        self.monotime.place(x=440, y=235)
        self.monotimereacell.place(x=500, y=260)
        for i in range(min(len(self.monoinitlistreacell), 4)):
            self.monoinitlistreacell[i].place(x=325, y=104 + i * 30)
            self.monoinitlistreac[i].place(x=325, y=115 + i * 30, anchor=tk.E)
        for i in range(min(len(self.monotermlistreacell), 4)):
            self.monotermlistreacell[i].place(x=325, y=414 + i * 30)
            self.monotermlistreac[i].place(x=325, y=425 + i * 30, anchor=tk.E)
        for i in range(min(len(self.monoproplistreacell), 4)):
            self.monoproplistreacell[i].place(x=325, y=259 + i * 30)
            self.monoproplistreac[i].place(x=325, y=270 + i * 30, anchor=tk.E)
        for i in range(min(len(self.monoconclistreacell), 5)):
            self.monoconclistreacell[i].place(x=670, y=74 + i * 30)
            self.monoconclistreac[i].place(x=670, y=85 + i * 30, anchor=tk.E)
        self.frame.place_forget()
        self.label.pack_forget()
        self.monoremc.place_forget()
        for i in self.monolistlabel:
            i.place_forget()

    def deletemono(self, mclick):
        if (40 - mclick.y) ** 2 + (40 - mclick.x) ** 2 < 400:
            if len(self.monolist) > 0:
                self.monoinitlistreacell[-1].place_forget()
                self.monoinitlistreac[-1].place_forget()
                self.monoconclistreacell[-1].place_forget()
                self.monoconclistreac[-1].place_forget()

                self.monolist.pop()
                self.monoconclistreac.pop()
                self.monoconclistreacell.pop()
                self.monoinitlistreac.pop()
                self.monoinitlistreacell.pop()
                for i in range(len(self.monolist) + (len(self.monolist) + 1)):
                    self.monoproplistreacell[-1].place_forget()
                    self.monoproplistreac[-1].place_forget()
                    self.monoproplistreac.pop()
                    self.monoproplistreacell.pop()
                    self.monoproplistfabric.pop()
                self.update_idletasks()

                x1 = 400
                x2 = 405
                y1 = 25
                if len(self.monoproplistreac) > 4:
                    y2 = 4 * 110 // len(self.monoproplistreac) + 25
                else:
                    y2 = 135
                r = 4
                points = (
                    x1 + r,
                    y1,
                    x1 + r,
                    y1,
                    x2 - r,
                    y1,
                    x2 - r,
                    y1,
                    x2,
                    y1,
                    x2,
                    y1 + r,
                    x2,
                    y1 + r,
                    x2,
                    y2 - r,
                    x2,
                    y2 - r,
                    x2,
                    y2,
                    x2 - r,
                    y2,
                    x2 - r,
                    y2,
                    x1 + r,
                    y2,
                    x1 + r,
                    y2,
                    x1,
                    y2,
                    x1,
                    y2 - r,
                    x1,
                    y2 - r,
                    x1,
                    y1 + r,
                    x1,
                    y1 + r,
                    x1,
                    y1,
                )
                self.monoprop.delete(self.polyprop)
                self.polyprop = self.monoprop.create_polygon(
                    points, smooth=True, fill="sea green"
                )

                y2 = min(135, 4 * 110 // len(self.monoinitlistreac) + 25)
                points = (
                    x1 + r,
                    y1,
                    x1 + r,
                    y1,
                    x2 - r,
                    y1,
                    x2 - r,
                    y1,
                    x2,
                    y1,
                    x2,
                    y1 + r,
                    x2,
                    y1 + r,
                    x2,
                    y2 - r,
                    x2,
                    y2 - r,
                    x2,
                    y2,
                    x2 - r,
                    y2,
                    x2 - r,
                    y2,
                    x1 + r,
                    y2,
                    x1 + r,
                    y2,
                    x1,
                    y2,
                    x1,
                    y2 - r,
                    x1,
                    y2 - r,
                    x1,
                    y1 + r,
                    x1,
                    y1 + r,
                    x1,
                    y1,
                )
                self.monoinit.delete(self.polyinit)
                self.polyinit = self.monoinit.create_polygon(
                    points, smooth=True, fill="sea green"
                )

                for i in range(
                    self.inititer,
                    min(len(self.monoinitlistreacell), 4) + self.inititer - 1,
                ):
                    self.monoinitlistreacell[i].place_forget()
                    self.monoinitlistreac[i].place_forget()

                for i in range(
                    self.propiter,
                    min(len(self.monoproplistreacell), 4)
                    + self.propiter
                    - len(self.monolist)
                    - (len(self.monolist) + 1),
                ):
                    self.monoproplistreacell[i].place_forget()
                    self.monoproplistreac[i].place_forget()

                for i in range(
                    self.conciter,
                    min(len(self.monoconclistreacell), 5) + self.conciter - 1,
                ):
                    self.monoconclistreacell[i].place_forget()
                    self.monoconclistreac[i].place_forget()

                self.propiter = 0
                self.inititer = 0
                self.conciter = 0

                for i in range(min(len(self.monoconclistreacell), 5)):
                    self.monoconclistreacell[i + self.conciter].place(
                        x=670, y=74 + i * 30
                    )
                    self.monoconclistreac[i + self.conciter].place(
                        x=670, y=85 + i * 30, anchor=tk.E
                    )

                for i in range(min(len(self.monoproplistreacell), 4)):
                    self.monoproplistreacell[i + self.propiter].place(
                        x=325, y=259 + i * 30
                    )
                    self.monoproplistreac[i + self.propiter].place(
                        x=325, y=270 + i * 30, anchor=tk.E
                    )

                for i in range(min(len(self.monoinitlistreacell), 4)):
                    self.monoinitlistreacell[i + self.inititer].place(
                        x=325, y=104 + i * 30
                    )
                    self.monoinitlistreac[i + self.inititer].place(
                        x=325, y=115 + i * 30, anchor=tk.E
                    )

                x1 = 335
                x2 = 340
                y2 = min(165, 5 * 140 // len(self.monoconclistreac) + 25)
                points = (
                    x1 + r,
                    y1,
                    x1 + r,
                    y1,
                    x2 - r,
                    y1,
                    x2 - r,
                    y1,
                    x2,
                    y1,
                    x2,
                    y1 + r,
                    x2,
                    y1 + r,
                    x2,
                    y2 - r,
                    x2,
                    y2 - r,
                    x2,
                    y2,
                    x2 - r,
                    y2,
                    x2 - r,
                    y2,
                    x1 + r,
                    y2,
                    x1 + r,
                    y2,
                    x1,
                    y2,
                    x1,
                    y2 - r,
                    x1,
                    y2 - r,
                    x1,
                    y1 + r,
                    x1,
                    y1 + r,
                    x1,
                    y1,
                )
                self.monocon.delete(self.polyconc)
                self.polyconc = self.monocon.create_polygon(
                    points, smooth=True, fill="sea green"
                )
            else:
                Message(
                    "il n'y a pas de monomère a supprimer",
                    (self.winfo_x(), self.winfo_y()),
                )

    def addmonof(self, mclick):
        if (40 - mclick.y) ** 2 + (40 - mclick.x) ** 2 < 400:
            Question(
                self, "Quel est le nom du monomère?", (self.winfo_x(), self.winfo_y())
            )

    def addmonof2(self, result):
        self.monoconclistreac.append(
            tk.Label(
                self,
                font=("Helvetica 15"),
                text=result + " :",
                bg="DarkSeaGreen1",
            )
        )
        self.monoconclistreacell.append(
            tk.Entry(
                self,
                width=7,
                font=("Helvetica 14"),
                background="DarkSeaGreen2",
            )
        )
        self.monoinitlistreac.append(
            tk.Label(
                self,
                font=("Helvetica 15"),
                text="I°+ " + result + " --> P" + result + "°",
                bg="DarkSeaGreen1",
            )
        )
        self.monoinitlistreacell.append(
            tk.Entry(
                self,
                width=7,
                font=("Helvetica 14"),
                background="DarkSeaGreen2",
            )
        )
        self.monoproplistreac.append(
            tk.Label(
                self,
                font=("Helvetica 15"),
                text="P" + result + "°+ " + result + "-->P" + result + "°",
                bg="DarkSeaGreen1",
            )
        )
        self.monoproplistreacell.append(
            tk.Entry(
                self,
                width=7,
                font=("Helvetica 14"),
                background="DarkSeaGreen2",
            )
        )

        self.monoproplistfabric.append((result, result))
        for i in self.monolist:
            self.monoproplistreac.append(
                tk.Label(
                    self,
                    font=("Helvetica 15"),
                    text="P" + i + "°+ " + result + "-->P" + result + "°",
                    bg="DarkSeaGreen1",
                )
            )
            self.monoproplistfabric.append((i, result))
            self.monoproplistreac.append(
                tk.Label(
                    self,
                    font=("Helvetica 15"),
                    text="P" + result + "°+ " + i + "-->P" + i + "°",
                    bg="DarkSeaGreen1",
                )
            )
            self.monoproplistfabric.append((result, i))
            self.monoproplistreacell.append(
                tk.Entry(
                    self,
                    width=7,
                    font=("Helvetica 14"),
                    background="DarkSeaGreen2",
                )
            )
            self.monoproplistreacell.append(
                tk.Entry(
                    self,
                    width=7,
                    font=("Helvetica 14"),
                    background="DarkSeaGreen2",
                )
            )

        self.monolist.append(result)

        for i in range(
            self.conciter, min(len(self.monoconclistreacell), 5) + self.conciter
        ):
            self.monoconclistreacell[i].place(x=670, y=74 + i * 30)
            self.monoconclistreac[i].place(x=670, y=85 + i * 30, anchor=tk.E)
        self.update_idletasks()

        for i in range(min(len(self.monoinitlistreacell), 4)):
            self.monoinitlistreacell[i + self.inititer].place(x=325, y=104 + i * 30)
            self.monoinitlistreac[i + self.inititer].place(
                x=325, y=115 + i * 30, anchor=tk.E
            )
        x1 = 400
        x2 = 405
        y1 = 25 + self.inititer * 110 // len(self.monoinitlistreac)
        r = 4
        y2 = min(
            135,
            4 * 110 // len(self.monoinitlistreac)
            + self.inititer * 110 // len(self.monoinitlistreac)
            + 25,
        )
        points = (
            x1 + r,
            y1,
            x1 + r,
            y1,
            x2 - r,
            y1,
            x2 - r,
            y1,
            x2,
            y1,
            x2,
            y1 + r,
            x2,
            y1 + r,
            x2,
            y2 - r,
            x2,
            y2 - r,
            x2,
            y2,
            x2 - r,
            y2,
            x2 - r,
            y2,
            x1 + r,
            y2,
            x1 + r,
            y2,
            x1,
            y2,
            x1,
            y2 - r,
            x1,
            y2 - r,
            x1,
            y1 + r,
            x1,
            y1 + r,
            x1,
            y1,
        )

        self.monoinit.delete(self.polyinit)
        self.polyinit = self.monoinit.create_polygon(
            points, smooth=True, fill="sea green"
        )

        for i in range(min(len(self.monoproplistreacell), 4)):
            self.monoproplistreacell[i + self.propiter].place(x=325, y=259 + i * 30)
            self.monoproplistreac[i + self.propiter].place(
                x=325, y=270 + i * 30, anchor=tk.E
            )

        y1 = 25 + self.propiter * 110 // len(self.monoproplistreac)
        y2 = min(
            135,
            4 * 110 // len(self.monoproplistreac)
            + self.propiter * 110 // len(self.monoproplistreac)
            + 25,
        )
        points = (
            x1 + r,
            y1,
            x1 + r,
            y1,
            x2 - r,
            y1,
            x2 - r,
            y1,
            x2,
            y1,
            x2,
            y1 + r,
            x2,
            y1 + r,
            x2,
            y2 - r,
            x2,
            y2 - r,
            x2,
            y2,
            x2 - r,
            y2,
            x2 - r,
            y2,
            x1 + r,
            y2,
            x1 + r,
            y2,
            x1,
            y2,
            x1,
            y2 - r,
            x1,
            y2 - r,
            x1,
            y1 + r,
            x1,
            y1 + r,
            x1,
            y1,
        )

        self.monoprop.delete(self.polyprop)
        self.polyprop = self.monoprop.create_polygon(
            points, smooth=True, fill="sea green"
        )

        for i in range(min(len(self.monoconclistreacell), 5)):
            self.monoconclistreacell[i + self.conciter].place(x=670, y=74 + i * 30)
            self.monoconclistreac[i + self.conciter].place(
                x=670, y=85 + i * 30, anchor=tk.E
            )
        x1 = 335
        x2 = 340
        y1 = 25 + self.conciter * 140 // len(self.monoconclistreac)
        y2 = min(
            165,
            5 * 140 // len(self.monoconclistreac)
            + self.conciter * 140 // len(self.monoconclistreac)
            + 25,
        )
        points = (
            x1 + r,
            y1,
            x1 + r,
            y1,
            x2 - r,
            y1,
            x2 - r,
            y1,
            x2,
            y1,
            x2,
            y1 + r,
            x2,
            y1 + r,
            x2,
            y2 - r,
            x2,
            y2 - r,
            x2,
            y2,
            x2 - r,
            y2,
            x2 - r,
            y2,
            x1 + r,
            y2,
            x1 + r,
            y2,
            x1,
            y2,
            x1,
            y2 - r,
            x1,
            y2 - r,
            x1,
            y1 + r,
            x1,
            y1 + r,
            x1,
            y1,
        )

        self.monocon.delete(self.polyconc)
        self.polyconc = self.monocon.create_polygon(
            points, smooth=True, fill="sea green"
        )

    def curseurmc(self, curs, mclick):
        if curs == "0":
            if (
                390 < mclick.x < 410
                and 25 < mclick.y < 135
                and len(self.monoinitlistreac) > 4
            ):
                newinit = max(
                    0, (mclick.y - 25) // (110 // len(self.monoinitlistreac)) - 3
                )
                if newinit != self.inititer:
                    for i in range(
                        self.inititer,
                        min(len(self.monoinitlistreacell), 4) + self.inititer,
                    ):
                        self.monoinitlistreacell[i].place_forget()
                        self.monoinitlistreac[i].place_forget()
                    self.inititer = min(newinit, len(self.monoinitlistreacell) - 4)
                    for i in range(min(len(self.monoinitlistreacell), 4)):
                        self.monoinitlistreacell[i + self.inititer].place(
                            x=325, y=104 + i * 30
                        )
                        self.monoinitlistreac[i + self.inititer].place(
                            x=325, y=115 + i * 30, anchor=tk.E
                        )
                    x1 = 400
                    x2 = 405
                    y1 = 25 + self.inititer * 110 // len(self.monoinitlistreac)
                    r = 4
                    y2 = min(
                        135,
                        4 * 110 // len(self.monoinitlistreac)
                        + self.inititer * 110 // len(self.monoinitlistreac)
                        + 25,
                    )
                    points = (
                        x1 + r,
                        y1,
                        x1 + r,
                        y1,
                        x2 - r,
                        y1,
                        x2 - r,
                        y1,
                        x2,
                        y1,
                        x2,
                        y1 + r,
                        x2,
                        y1 + r,
                        x2,
                        y2 - r,
                        x2,
                        y2 - r,
                        x2,
                        y2,
                        x2 - r,
                        y2,
                        x2 - r,
                        y2,
                        x1 + r,
                        y2,
                        x1 + r,
                        y2,
                        x1,
                        y2,
                        x1,
                        y2 - r,
                        x1,
                        y2 - r,
                        x1,
                        y1 + r,
                        x1,
                        y1 + r,
                        x1,
                        y1,
                    )

                    self.monoinit.delete(self.polyinit)
                    self.polyinit = self.monoinit.create_polygon(
                        points, smooth=True, fill="sea green"
                    )

        if curs == "1":
            if (
                390 < mclick.x < 410
                and 25 < mclick.y < 135
                and len(self.monoproplistreac) > 4
            ):
                newprop = max(
                    0, (mclick.y - 25) // (110 // len(self.monoproplistreac)) - 3
                )
                if newprop != self.propiter:
                    for i in range(
                        self.propiter,
                        min(len(self.monoproplistreacell), 4) + self.propiter,
                    ):
                        self.monoproplistreacell[i].place_forget()
                        self.monoproplistreac[i].place_forget()
                    self.propiter = min(newprop, len(self.monoproplistreacell) - 4)
                    for i in range(min(len(self.monoproplistreacell), 4)):
                        self.monoproplistreacell[i + self.propiter].place(
                            x=325, y=259 + i * 30
                        )
                        self.monoproplistreac[i + self.propiter].place(
                            x=325, y=270 + i * 30, anchor=tk.E
                        )
                    x1 = 400
                    x2 = 405
                    y1 = 25 + self.propiter * 110 // len(self.monoproplistreac)
                    r = 4
                    y2 = min(
                        135,
                        4 * 110 // len(self.monoproplistreac)
                        + self.propiter * 110 // len(self.monoproplistreac)
                        + 25,
                    )
                    points = (
                        x1 + r,
                        y1,
                        x1 + r,
                        y1,
                        x2 - r,
                        y1,
                        x2 - r,
                        y1,
                        x2,
                        y1,
                        x2,
                        y1 + r,
                        x2,
                        y1 + r,
                        x2,
                        y2 - r,
                        x2,
                        y2 - r,
                        x2,
                        y2,
                        x2 - r,
                        y2,
                        x2 - r,
                        y2,
                        x1 + r,
                        y2,
                        x1 + r,
                        y2,
                        x1,
                        y2,
                        x1,
                        y2 - r,
                        x1,
                        y2 - r,
                        x1,
                        y1 + r,
                        x1,
                        y1 + r,
                        x1,
                        y1,
                    )

                    self.monoprop.delete(self.polyprop)
                    self.polyprop = self.monoprop.create_polygon(
                        points, smooth=True, fill="sea green"
                    )

        if curs == "2":
            if (
                320 < mclick.x < 350
                and 25 < mclick.y < 165
                and len(self.monoconclistreac) > 5
            ):
                newconc = max(
                    0, (mclick.y - 25) // (140 // len(self.monoconclistreac)) - 4
                )
                if newconc != self.conciter:
                    for i in range(
                        self.conciter,
                        min(len(self.monoconclistreacell), 5) + self.conciter,
                    ):
                        self.monoconclistreacell[i].place_forget()
                        self.monoconclistreac[i].place_forget()
                    self.conciter = min(newconc, len(self.monoconclistreacell) - 5)
                    for i in range(min(len(self.monoconclistreacell), 5)):
                        self.monoconclistreacell[i + self.conciter].place(
                            x=670, y=74 + i * 30
                        )
                        self.monoconclistreac[i + self.conciter].place(
                            x=670, y=85 + i * 30, anchor=tk.E
                        )
                    x1 = 335
                    x2 = 340
                    y1 = 25 + self.conciter * 140 // len(self.monoconclistreac)
                    r = 4
                    y2 = min(
                        165,
                        5 * 140 // len(self.monoconclistreac)
                        + self.conciter * 140 // len(self.monoconclistreac)
                        + 25,
                    )
                    points = (
                        x1 + r,
                        y1,
                        x1 + r,
                        y1,
                        x2 - r,
                        y1,
                        x2 - r,
                        y1,
                        x2,
                        y1,
                        x2,
                        y1 + r,
                        x2,
                        y1 + r,
                        x2,
                        y2 - r,
                        x2,
                        y2 - r,
                        x2,
                        y2,
                        x2 - r,
                        y2,
                        x2 - r,
                        y2,
                        x1 + r,
                        y2,
                        x1 + r,
                        y2,
                        x1,
                        y2,
                        x1,
                        y2 - r,
                        x1,
                        y2 - r,
                        x1,
                        y1 + r,
                        x1,
                        y1 + r,
                        x1,
                        y1,
                    )

                    self.monocon.delete(self.polyconc)
                    self.polyconc = self.monocon.create_polygon(
                        points, smooth=True, fill="sea green"
                    )

    def simulationmc(self):
        try:
            # c'est ici que les choses intéréssantes commencent
            if len(self.monolist) > 0:
                tempost = float(self.monotimereacell.get())
                dicoinit = {}
                dicoinit["I-I"] = (
                    float(self.monoinitlistreacell[0].get()),
                    float(self.monoconclistreacell[0].get()),
                )
                dicoinit["I°"] = (
                    float(self.monoinitlistreacell[1].get()),
                    float(self.monoconclistreacell[1].get()),
                )
                tupletermine = (
                    float(self.monotermlistreacell[0].get()),
                    float(self.monotermlistreacell[1].get()),
                )
                for i in range(len(self.monolist)):
                    dicoinit[self.monolist[i]] = (
                        float(self.monoinitlistreacell[2 + i].get()),
                        float(self.monoconclistreacell[2 + i].get()),
                    )

                listeprop = []
                for i in range(len(self.monoproplistfabric)):
                    listeprop.append(
                        (
                            self.monoproplistfabric[i],
                            float(self.monoproplistreacell[i].get()),
                        )
                    )
                self.listelongueur = simulmc(
                    int(self.limchaine.get()),
                    dicoinit,
                    listeprop,
                    tupletermine,
                    tempost,
                )
                self.affichemc()
            else:
                Message(
                    "la simulation nécessite au moins un monomère",
                    (self.winfo_x(), self.winfo_y()),
                )
        except ZeroDivisionError:
            Message(
                [
                    "erreur de calcul",
                    "sources d'erreurs possibles:",
                    "-la simulation entraîne une division par 0,",
                    " essayez de nouvelles valeurs",
                ],
                (self.winfo_x(), self.winfo_y()),
            )
        except ValueError:
            Message(
                [
                    "erreur de synthaxe",
                    "sources d'erreurs possibles:",
                    "-vous avez peut-être oublié de remplir une case.",
                    "-vous avez peut-être confondu un point et une virgule:",
                    "4,2 -> synthaxe correcte: 4.2 (ne tapez jamais de virgules !!)",
                    "-vous avez peut-être mal écrit une forme exponentielle:",
                    "5.3*10^-5 -> synthaxe correcte: 5.3e-5",
                ],
                (self.winfo_x(), self.winfo_y()),
            )

    def affichemc(self):
        self.situation = "affichemc"
        self.addmono.place_forget()
        self.limchaine.place_forget()
        self.menuderoul.place_forget()
        self.format.place_forget()
        self.chemintext.place_forget()
        self.chemin.place_forget()
        self.cheminbouton.place_forget()
        self.nom.place_forget()
        self.nomtabl.place_forget()
        self.enregistrebouton.place_forget()

        self.monolabel.place_forget()
        self.mcconfirm.place_forget()
        self.erasemono.place_forget()
        self.monocon.place_forget()
        self.framemc.place(x=100, y=0)
        self.labelmc.pack()
        self.rapportcanv1.place(x=0, y=0)

        self.monoprop.place_forget()
        self.monoinit.place_forget()
        self.monotermination.place_forget()
        self.monotime.place_forget()
        self.monotimereacell.place_forget()
        for i in range(min(len(self.monoinitlistreacell), 4)):
            self.monoinitlistreacell[i].place_forget()
            self.monoinitlistreac[i].place_forget()
        for i in range(min(len(self.monotermlistreacell), 4)):
            self.monotermlistreacell[i].place_forget()
            self.monotermlistreac[i].place_forget()
        for i in range(min(len(self.monoproplistreacell), 4)):
            self.monoproplistreacell[i].place_forget()
            self.monoproplistreac[i].place_forget()
        for i in range(min(len(self.monoconclistreacell), 5)):
            self.monoconclistreacell[i].place_forget()
            self.monoconclistreac[i].place_forget()

        self.simulmc = MovableImage(self, "simulmc.bmp")
        self.simulmc2 = MovableImage(self, "simulmc2.bmp")
        self.simulmc.place(x=525, y=60)
        self.simulmc2.place(x=525, y=280)
        for i in self.monolistlabel:
            i.place_forget()
        self.monolistlabel = []
        self.monoremc.place(x=40, y=120)
        j = 0
        k = 2
        for i in self.listelongueur[1]:
            e = i[0]
            n = round(i[1], 3)
            self.monolistlabel.append(
                tk.Label(
                    self,
                    font=("Helvetica 15"),
                    text=e + ": " + str(n),
                    bg="DarkSeaGreen2",
                    relief=tk.SUNKEN,
                    foreground=self.listelongueur[3][k][1],
                )
            )
            self.monolistlabel[-1].place(x=50, y=170 + j * 30)
            j += 1
            k += 1
        self.distribmc.place(x=100, y=440)
        self.radiomc.place(x=100, y=410)
        if self.enrv == 1:
            self.enrv = 0
            fini = Message(
                "le document est enregistré.", (self.winfo_x(), self.winfo_y())
            )

    def distriblanceoublie(self):
        try:
            self.widget4.place_forget()
        except AttributeError:
            pass
        self.monoremc.place(x=40, y=120)
        j = 0
        for i in self.monolistlabel:
            i.place(x=50, y=170 + j * 30)
            j += 1

    def distribmclance(self):
        try:
            self.widget4.place_forget()
        except AttributeError:
            pass

        x = {}
        y = range(100)
        for key in self.listelongueur[2]:
            if key != "I-I" and key != "I°":
                x[key] = [0, 0, 0, 0, 0] * 20
                for i in self.listelongueur[2][key]:
                    if i < 100:
                        x[key][i] += 1

        self.figmc = plt.figure(figsize=(4, 4), dpi=75)
        bx = self.figmc.add_subplot(111)
        j = 0
        for key in x:
            for k in range(len(x[key])):
                tex = x[key][k]
                x[key][k] = tex * k
            dd1 = bx.plot(y, x[key], color=self.listelongueur[3][j + 2][2], label=key)

            j += 1
        bx.set_xscale("log")

        self.figmc.legend()

        canvas3 = FigureCanvasTkAgg(self.figmc, master=self)  # A tk.DrawingArea.
        canvas3.draw()
        canvas3.mpl_connect(
            "key_press_event", lambda event: print(f"you pressed {event.key}")
        )
        canvas3.mpl_connect("key_press_event", key_press_handler)
        self.widget4 = canvas3.get_tk_widget()
        self.widget4.place(x=30, y=80)
        self.monoremc.place_forget()
        for i in self.monolistlabel:
            i.place_forget()

    #################################################################
    #                                                               #
    #                simulation analytique                          #
    #                                                               #
    #################################################################
    def simul1(self):
        self.situation = "simul1"
        self.rapportbout.pack_forget()
        self.analbout.place_forget()
        self.mcbout.place_forget()
        self.canvmenu.pack_forget()
        self.rapportcanv1.place(x=0, y=0)
        self.rapportcanv22.place(x=0, y=80)
        self.simulconfirme11.place(x=100, y=400)
        self.simulconfirme12.place(x=500, y=400)
        self.k11cell.place(x=150, y=150)
        self.k22cell.place(x=300, y=200)
        self.k21cell.place(x=150, y=200)
        self.k12cell.place(x=300, y=150)
        self.M1cell.place(x=150, y=250)
        self.M2cell.place(x=300, y=250)
        self.f1cell.place(x=525, y=255)
        self.r1cell.place(x=525, y=155)
        self.r2cell.place(x=675, y=155)
        self.f1lab.place(x=480, y=250)
        self.r1lab.place(x=480, y=150)
        self.r2lab.place(x=630, y=150)
        self.M1lab.place(x=80, y=250)
        self.M2lab.place(x=230, y=250)
        self.k11lab.place(x=80, y=150)
        self.k22lab.place(x=230, y=200)
        self.k21lab.place(x=80, y=200)
        self.k12lab.place(x=230, y=150)
        self.simullance.place_forget()
        self.distriblance.place_forget()
        self.resultsim.place_forget()
        self.result.place_forget()
        self.frame.place_forget()
        self.label.pack_forget()
        self.reparta.place_forget()
        self.repartb.place_forget()
        self.repartlabcanv.place_forget()
        self.repartlabcanv.delete(tk.ALL)

    def simul21(self):
        try:
            m1 = float(self.M1cell.get())
            m2 = float(self.M2cell.get())
            k11 = float(self.k11cell.get())
            k12 = float(self.k12cell.get())
            k21 = float(self.k21cell.get())
            k22 = float(self.k22cell.get())

            self.r1 = k11 / k12
            self.r2 = k22 / k21
            self.f10 = m1 / (m1 + m2)
            self.p11 = k11 * m1 / (k11 * m1 + k12 * m2)
            self.p22 = k22 * m2 / (k22 * m2 + k21 * m1)
            tha = 1 / (1 - self.p11)
            thb = 1 / (1 - self.p22)
            self.result.config(
                text="taille théorique d'une chaine de composé 1:\n"
                + str(round(tha, 2))
                + "\ntaille théorique d'une chaine de composé 2:\n"
                + str(round(thb, 2))
            )

            self.situation = "simulaffiche"
            self.simullance.place(x=400, y=400)

            self.distriblance.place(x=100, y=400)
            self.rapportcanv22.place_forget()
            self.simulconfirme11.place_forget()
            self.simulconfirme12.place_forget()
            self.k11cell.place_forget()
            self.k12cell.place_forget()
            self.k21cell.place_forget()
            self.k22cell.place_forget()
            self.M1cell.place_forget()
            self.M2cell.place_forget()
            self.f1cell.place_forget()
            self.r1cell.place_forget()
            self.r2cell.place_forget()
            self.f1lab.place_forget()
            self.r1lab.place_forget()
            self.r2lab.place_forget()
            self.M1lab.place_forget()
            self.M2lab.place_forget()
            self.k11lab.place_forget()
            self.k12lab.place_forget()
            self.k21lab.place_forget()
            self.k22lab.place_forget()

            self.result.place(x=10, y=150)
        except ValueError:
            Message(
                [
                    "erreur de synthaxe",
                    "sources d'erreurs possibles:",
                    "-vous avez peut-être oublié de remplir une case.",
                    "-vous avez peut-être confondu un point et une virgule:",
                    "4,2 -> synthaxe correcte: 4.2 (ne tapez jamais de virgules !!)",
                    "-vous avez peut-être mal écrit une forme exponentielle:",
                    "5.3*10^-5 -> synthaxe correct: 5.3e-5",
                ],
                (self.winfo_x(), self.winfo_y()),
            )
        except ZeroDivisionError:
            Message(
                [
                    "erreur de calcul",
                    "sources d'erreurs possibles:",
                    "- la calcul avec ces données entraîne une division par 0",
                ],
                (self.winfo_x(), self.winfo_y()),
            )

    def simul22(self):
        try:
            f1 = float(self.f1cell.get())
            f2 = 1 - f1
            r1 = float(self.r1cell.get())
            r2 = float(self.r2cell.get())
            self.r1 = r1
            self.r2 = r2
            self.f10 = f1
            self.p11 = f1 / (f1 + f2 / r1)
            self.p22 = f2 / (f2 + f1 / r2)
            tha = 1 / (1 - self.p11)
            thb = 1 / (1 - self.p22)
            self.result.config(
                text="taille théorique d'une chaine de composé 1:\n"
                + str(round(tha, 2))
                + "\ntaille théorique d'une chaine de composé 2:\n"
                + str(round(thb, 2))
            )
            self.result.place(x=10, y=150)
            self.situation = "simulaffiche"
            self.rapportcanv22.place_forget()
            self.simullance.place(x=400, y=400)
            self.distriblance.place(x=100, y=400)
            self.simulconfirme11.place_forget()
            self.simulconfirme12.place_forget()
            self.k11cell.place_forget()
            self.k12cell.place_forget()
            self.k21cell.place_forget()
            self.k22cell.place_forget()
            self.M1cell.place_forget()
            self.M2cell.place_forget()
            self.f1cell.place_forget()
            self.r1cell.place_forget()
            self.r2cell.place_forget()
            self.f1lab.place_forget()
            self.r1lab.place_forget()
            self.r2lab.place_forget()
            self.M1lab.place_forget()
            self.M2lab.place_forget()
            self.k11lab.place_forget()
            self.k12lab.place_forget()
            self.k21lab.place_forget()
            self.k22lab.place_forget()

        except ValueError:
            Message(
                [
                    "erreur de synthaxe",
                    "sources d'erreurs possibles:",
                    "-vous avez peut-êtres oublié de remplir une case.",
                    "-vous avez peut-être confondu un point et une virgule:",
                    "4,2 -> synthaxe correcte: 4.2 (ne tapez jamais de virgules !!)",
                    "-vous avez peut-être mal écrit une forme exponentiel:",
                    "5.3*10^-5 -> synthaxe correcte: 5.3e-5",
                ],
                (self.winfo_x(), self.winfo_y()),
            )
        except ZeroDivisionError:
            Message(
                [
                    "erreur de calcul",
                    "sources d'erreurs possibles:",
                    "- le calcul avec ces données entraîne une division par 0",
                ],
                (self.winfo_x(), self.winfo_y()),
            )

    def simul3(self):
        simulres = simulanal(1000, self.p11, self.p22)
        tha = simulres[0]
        resa = simulres[1]
        thb = simulres[2]
        resb = simulres[3]
        mina = simulres[4]
        maxa = simulres[5]
        minb = simulres[6]
        maxb = simulres[7]
        arepart = simulres[8]
        brepart = simulres[9]
        ttime = simulres[10]
        self.resultsim.configure(
            text="taille moyenne theorique 1: "
            + str(round(tha, 4))
            + "\ntaille moyenne trouvée 1: "
            + str(round(resa, 4))
            + "\ntaille moyenne theorique 2: "
            + str(round(thb, 4))
            + "\ntaille moyenne trouvée 2: "
            + str(round(resb, 4))
            + "\nlongueur de 1 inclue entre "
            + str(round(mina, 4))
            + " et "
            + str(round(maxa, 4))
            + "\nlongueur de 2 inclue entre "
            + str(round(minb, 4))
            + " et "
            + str(round(maxb, 4))
            + "\ntemps:"
            + str(round(ttime, 4)),
        )

        self.reparta.place(x=50, y=400)
        self.repartb.place(x=450, y=400)
        couleur = [
            "orange",
            "magenta",
            "sienna",
            "silver",
            "aqua",
            "green",
            "yellow",
            "lightpink",
            "red",
            "royalblue",
        ]
        acouleur = {}
        couleuract = ""
        for i in range(len(arepart) - 1):
            acouleur[arepart[i]] = couleur[i]
        for i in range(len(arepart) - 1):
            if couleuract != acouleur[arepart[i]]:
                self.repartlabcanv.create_text(
                    50 + 30 * i, 15, font=("Helvetica 13 bold"), text=arepart[i]
                )
            couleuract = acouleur[arepart[i]]
            self.reparta.create_rectangle(
                2 + 30 * i, 0, 2 + 30 * (i + 1), 22, fill=acouleur[arepart[i]], width=0
            )
        bcouleur = {}
        for i in range(len(brepart) - 1):
            bcouleur[brepart[i]] = couleur[i]
        couleuract = ""
        for i in range(len(brepart) - 1):
            if couleuract != bcouleur[brepart[i]]:
                self.repartlabcanv.create_text(
                    450 + 30 * i, 15, font=("Helvetica 13 bold"), text=brepart[i]
                )
            couleuract = bcouleur[brepart[i]]
            self.repartb.create_rectangle(
                2 + 30 * i, 0, 2 + 30 * (i + 1), 22, fill=bcouleur[brepart[i]], width=0
            )
        self.repartlabcanv.create_text(
            350, 15, font=("Helvetica 13 bold"), text=arepart[-1]
        )
        self.repartlabcanv.create_text(
            750, 15, font=("Helvetica 13 bold"), text=brepart[-1]
        )
        self.repartlabcanv.place(x=0, y=425)
        self.resultsim.place(x=20, y=100)
        self.simullance.place_forget()
        self.distriblance.place_forget()
        self.result.place_forget()

    def simuldist(self):
        try:
            lambdas = range(20)

            self.figa = plt.figure(figsize=(6, 6), dpi=75)

            self.figa2 = plt.figure(figsize=(4, 4), dpi=75)

            self.figa.patch.set_facecolor((144 / 255, 238 / 255, 144 / 255))
            self.figa2.patch.set_facecolor((144 / 255, 238 / 255, 144 / 255))

            ax1 = self.figa.add_subplot(111, projection="3d")
            ax1.set_facecolor((144 / 255, 238 / 255, 144 / 255))

            bx = self.figa2.add_subplot(111)

            x, y = distribanal(self.r1, self.r2, self.f10)

            verts = [polygon_under_graph(np.log10(x), y[l]) for l in lambdas]

            lambdas = np.arange(20) % 10

            facecolors = plt.colormaps["hsv"](np.linspace(0, 1, len(verts)))
            # You still have to take log10(Z) but thats just one operation
            poly = PolyCollection(
                verts, facecolors=facecolors, edgecolor="black", alpha=0.7
            )
            ysuprem1 = sum(y[:10]) / len(y[:10])
            ysuprem2 = sum(y[10:]) / len(y[10:])

            ax1.add_collection3d(poly, zs=lambdas, zdir="y")
            ax1.xaxis.set_major_formatter(mticker.FuncFormatter(log_tick_formatter))
            ax1.xaxis.set_major_locator(mticker.MaxNLocator(integer=True))
            ax1.yaxis.set_major_formatter(mticker.FuncFormatter(dizaine_formatter))
            ax1.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
            ax1.set_xlim(0, 3)
            ax1.set_ylim(0, 9)
            ax1.set_xlabel("taille de la chaîne de la molécule")
            ax1.set_ylabel("conversion")
            ax1.set_zlabel("probabilité normalisée")
            ax1.view_init(elev=10, azim=-80)
            dd1 = bx.plot(x, y[0])
            dd2 = bx.plot(x, y[10])
            bx.set_xlabel("taille de la chaîne de la molécule")
            bx.set_ylabel("probabilité normalisée")
            bx.set_xscale("log")
            canvas = FigureCanvasTkAgg(self.figa, master=self)  # A tk.DrawingArea.
            canvas.draw()
            canvas.mpl_connect(
                "key_press_event", lambda event: print(f"you pressed {event.key}")
            )
            canvas.mpl_connect("key_press_event", key_press_handler)
            self.widget2 = canvas.get_tk_widget()
            self.widget2.postscript(
                file="testmatplot.ps", x=0, y=0, height=300, width=450
            )

            self.widget2.place(x=320, y=00)

            canvas2 = FigureCanvasTkAgg(self.figa2, master=self)  # A tk.DrawingArea.
            canvas2.draw()
            canvas2.mpl_connect(
                "key_press_event", lambda event: print(f"you pressed {event.key}")
            )
            canvas2.mpl_connect("key_press_event", key_press_handler)
            self.widget3 = canvas2.get_tk_widget()
            self.widget3.place(x=30, y=80)
            self.cursor = tk.Scale(
                self,
                from_=0,
                to=90,
                resolution=10,
                foreground="black",
                orient="horizontal",
                command=lambda data: [
                    dd1[0].set_data(x, y[int(data) // 10]),
                    dd2[0].set_data(x, y[10 + int(data) // 10]),
                    canvas2.draw(),
                ],
                bg="light green",
                highlightthickness=0,
                troughcolor="white",
                activebackground="light grey",
                font="Helvetica 25 bold",
                length=150,
                width=20,
            )
            self.variable = tk.StringVar()
            self.variable.set(value="1")
            self.radio1 = tk.Radiobutton(
                self,
                variable=self.variable,
                value="1",
                text="conversion donnée",
                font="Helvetica 15 bold",
                activebackground="DarkSeaGreen2",
                background="DarkSeaGreen1",
                highlightbackground="darkblue",
                command=lambda: [
                    dd1[0].set_data(x, y[int(self.cursor.get()) // 10]),
                    dd2[0].set_data(x, y[10 + int(self.cursor.get()) // 10]),
                    canvas2.draw(),
                    self.cursor.place(x=100, y=50),
                ],
            )
            self.radio2 = tk.Radiobutton(
                self,
                variable=self.variable,
                value="2",
                text="moyenne",
                font="Helvetica 15 bold",
                activebackground="DarkSeaGreen2",
                background="DarkSeaGreen1",
                highlightbackground="darkblue",
                command=lambda: [
                    dd1[0].set_data(x, ysuprem1),
                    dd2[0].set_data(x, ysuprem2),
                    self.cursor.place_forget(),
                    canvas2.draw(),
                ],
            )
            self.cursor.place(x=100, y=50)
            self.radio1.place(x=100, y=400)
            self.radio2.place(x=100, y=430)
            self.frameanal1.place(x=00, y=400)
            self.labelanal1.pack()
            self.situation = "simuldistrib"
            self.simullance.place_forget()
            self.menuderoul.place_forget()
            self.format.place_forget()
            self.chemintext.place_forget()
            self.chemin.place_forget()
            self.cheminbouton.place_forget()
            self.nom.place_forget()
            self.nomtabl.place_forget()
            self.enregistrebouton.place_forget()
            self.distriblance.place_forget()
            self.result.place_forget()
            if self.enrv == 1:
                self.enrv = 0
                fini = Message(
                    "le document est enregistré", (self.winfo_x(), self.winfo_y())
                )

        except ZeroDivisionError:
            Message(
                [
                    "erreur de calcul",
                    "sources d'erreurs possibles:",
                    "- la calcul avec ces données entraîne une division par 0",
                ],
                (self.winfo_x(), self.winfo_y()),
            )

    #################################################################
    #                                                               #
    #                rapport de reaction                            #
    #                                                               #
    #################################################################

    def checkpreccurseurt(self, mclick):
        if (mclick.x - self.placecurprec - 20) ** 2 + (mclick.y - 20) ** 2 < 400:
            self.curseurselect = "prec"

    def freepreccurseurt(self, mclick):
        self.curseurselect = ""

    def updatepreccurseurt(self, mclick):
        if self.curseurselect == "prec":
            if mclick.x - 20 < 0:
                self.precslider.moveto(self.tranchcurseurprec, 0)
                self.placecurprec = 0
            elif mclick.x - 20 > 180:
                self.precslider.moveto(self.tranchcurseurprec, 180)
                self.placecurprec = 180
            else:
                self.precslider.moveto(self.tranchcurseurprec, mclick.x - 20)
                self.placecurprec = mclick.x - 20
            if self.placecurprec < 90:
                self.preclabval = round(1 - self.placecurprec / 100, 2)
            else:
                self.preclabval = round(0.1 - (self.placecurprec - 90) / 1000, 2)

            self.preclab.config(text=str(self.preclabval))
        self.precslider.coords(
            self.sliderpreclignep,
            [20, 20, self.placecurprec + 20, 20],
        )
        temps = (
            (self.preclabval) ** -2
            * (self.fintranchpot - self.debtranchpot)
            * (self.fintranchs - self.debtranchs)
            / 9000
        )
        min = int(temps // 60)
        s = round((temps % 60) // 1)
        cent = round((temps % 1) * 100)
        self.estimation.config(
            text="estimation du temps de calcul: \n"
            + str(min)
            + "min "
            + str(s)
            + "s "
            + str(cent)
        )

    def checkpotcurseurt(self, mclick):
        if (mclick.x - self.placecurpot1 - 20) ** 2 + (mclick.y - 20) ** 2 < 400:
            self.curseurselect = "pot1"

        elif (mclick.x - self.placecurpot2 - 20) ** 2 + (mclick.y - 20) ** 2 < 400:
            self.curseurselect = "pot2"

    def freepotcurseurt(self, mclick):
        self.curseurselect = ""

    def updatepotcurseurt(self, mclick):
        if self.curseurselect == "pot1":
            if mclick.x - 20 < 0:
                self.tranchsliderpot.moveto(self.tranchcurseurpot1, 0)
                self.placecurs1 = 0
            elif mclick.x - 20 > self.placecurpot2 - 40:
                self.tranchsliderpot.moveto(
                    self.tranchcurseurpot1, self.placecurpot2 - 40
                )
                self.placecurpot1 = self.placecurpot2 - 40
            else:
                self.tranchsliderpot.moveto(self.tranchcurseurpot1, mclick.x - 20)
                self.placecurpot1 = mclick.x - 20
            if self.placecurpot1 < 150:
                self.debtranchpot = round(self.placecurpot1 / 50, 2)
            elif self.placecurs1 < 290:
                self.debtranchpot = round(3 + (self.placecurpot1 - 150) / 20, 2)
            else:
                self.debtranchpot = round(10 + (self.placecurpot1 - 290) / 2, 2)

            self.tranchepot1lab.config(text=self.debtranchpot)
            self.tranchepot1lab.place_configure(x=self.placecurpot1 + 350)
        if self.curseurselect == "pot2":
            if mclick.x - 20 > 400:
                self.tranchsliderpot.moveto(self.tranchcurseurpot2, 400)
                self.placecurpot2 = 400
            elif mclick.x - 20 < self.placecurpot1 + 40:
                self.tranchsliderpot.moveto(
                    self.tranchcurseurpot2, self.placecurpot1 + 40
                )
                self.placecurpot2 = self.placecurpot1 + 40

            else:
                self.tranchsliderpot.moveto(self.tranchcurseurpot2, mclick.x - 20)
                self.placecurpot2 = mclick.x - 20
            if self.placecurpot2 < 150:
                self.fintranchpot = round(self.placecurpot2 / 50, 2)
            elif self.placecurpot2 < 290:
                self.fintranchpot = round(3 + (self.placecurpot2 - 150) / 20, 2)
            else:
                self.fintranchpot = round(10 + (self.placecurpot2 - 290) / 2, 2)

            self.tranchepot2lab.config(text=self.fintranchpot)
            self.tranchepot2lab.place_configure(x=self.placecurpot2 + 350)
        self.tranchsliderpot.coords(
            self.sliderpotlignep,
            [self.placecurpot1 + 20, 20, self.placecurpot2 + 20, 20],
        )
        temps = (
            (self.preclabval) ** -2
            * (self.fintranchpot - self.debtranchpot)
            * (self.fintranchs - self.debtranchs)
            / 9000
        )
        min = int(temps // 60)
        s = round((temps % 60) // 1)
        cent = round((temps % 1) * 100)
        self.estimation.config(
            text="estimation du temps de calcul: \n"
            + str(min)
            + "min "
            + str(s)
            + "s "
            + str(cent)
        )

    def checkscurseurt(self, mclick):
        if (mclick.x - self.placecurs1 - 20) ** 2 + (mclick.y - 20) ** 2 < 400:
            self.curseurselect = "s1"

        elif (mclick.x - self.placecurs2 - 20) ** 2 + (mclick.y - 20) ** 2 < 400:
            self.curseurselect = "s2"

    def freescurseurt(self, mclick):
        self.curseurselect = ""

    def updatescurseurt(self, mclick):
        if self.curseurselect == "s1":
            if mclick.x - 20 < 0:
                self.tranchsliders.moveto(self.tranchcurseurs1, 0)
                self.placecurs1 = 0
            elif mclick.x - 20 > self.placecurs2 - 40:
                self.tranchsliders.moveto(self.tranchcurseurs1, self.placecurs2 - 40)
                self.placecurs1 = self.placecurs2 - 40
            else:
                self.tranchsliders.moveto(self.tranchcurseurs1, mclick.x - 20)
                self.placecurs1 = mclick.x - 20
            if self.placecurs1 < 150:
                self.debtranchs = round(self.placecurs1 / 50, 2)
            elif self.placecurs1 < 290:
                self.debtranchs = round(3 + (self.placecurs1 - 150) / 20, 2)
            else:
                self.debtranchs = round(10 + (self.placecurs1 - 290) / 2, 2)

            self.tranches1lab.config(text=self.debtranchs)
            self.tranches1lab.place_configure(x=self.placecurs1 + 350)
        if self.curseurselect == "s2":
            if mclick.x - 20 > 400:
                self.tranchsliders.moveto(self.tranchcurseurs2, 400)
                self.placecurs2 = 400
            elif mclick.x - 20 < self.placecurs1 + 40:
                self.tranchsliders.moveto(self.tranchcurseurs2, self.placecurs1 + 40)
                self.placecurs2 = self.placecurs1 + 40

            else:
                self.tranchsliders.moveto(self.tranchcurseurs2, mclick.x - 20)
                self.placecurs2 = mclick.x - 20
            if self.placecurs2 < 150:
                self.fintranchs = round(self.placecurs2 / 50, 2)
            elif self.placecurs2 < 290:
                self.fintranchs = round(3 + (self.placecurs2 - 150) / 20, 2)
            else:
                self.fintranchs = round(10 + (self.placecurs2 - 290) / 2, 2)

            self.tranches2lab.config(text=self.fintranchs)
            self.tranches2lab.place_configure(x=self.placecurs2 + 350)

        self.tranchsliders.coords(
            self.sliderslignep, [self.placecurs1 + 20, 20, self.placecurs2 + 20, 20]
        )
        temps = (
            (self.preclabval) ** -2
            * (self.fintranchpot - self.debtranchpot)
            * (self.fintranchs - self.debtranchs)
            / 9000
        )
        min = int(temps // 60)
        s = round((temps % 60) // 1)
        cent = round((temps % 1) * 100)
        self.estimation.config(
            text="estimation du temps de calcul: \n"
            + str(min)
            + "min "
            + str(s)
            + "s "
            + str(cent)
        )

    def browse_button(self):
        filename = filedialog.askdirectory()
        self.chemin.insert(tk.END, filename)

    def browse_button2(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.cheminex.delete(0, tk.END)
        self.cheminex.insert(tk.END, filename)

    def rapport1(self):
        self.situation = "rapport1"
        self.preclabel.place_forget()
        self.tranchslabel.place_forget()
        self.tranches1lab.place_forget()
        self.tranches2lab.place_forget()
        self.tranchepot1lab.place_forget()
        self.tranchepot2lab.place_forget()
        self.tranchpotlabel.place_forget()
        self.rapportbout.pack_forget()
        self.analbout.place_forget()
        self.mcbout.place_forget()
        self.canvmenu.pack_forget()
        self.rapportcanv1.place(x=0, y=0)
        self.rapportcanv12.place(x=0, y=80)
        self.manuel.place(x=60, y=100)
        self.experiencetexte.place(x=60, y=190)
        self.experiencecurseur.place(x=100, y=266)
        self.importation.place(x=460, y=90)
        self.rapportconfirme1.place(x=100, y=400)
        self.rapportconfirme2.place(x=500, y=400)
        self.cheminex.place(x=420, y=266)
        self.cheminboutonex.place(x=420, y=310)
        self.cheminextext.place(x=460, y=210)
        self.frame3.place(x=700, y=0)
        self.label3.place(x=0, y=0)
        self.frame.place_forget()
        self.label.pack_forget()
        self.exp[0].pack_forget()
        self.tranchsliders.place_forget()
        self.tranchsliderpot.place_forget()
        self.precslider.place_forget()
        self.preclab.place_forget()
        self.estimation.place_forget()
        self.composélabel.place(x=120, y=10)
        self.composéAcell.place(x=370, y=15)
        self.composéBcell.place(x=430, y=15)

    def parametre(self, mclick):
        self.comps = self.composéAcell.get()
        self.comppot = self.composéBcell.get()
        self.tranchslabel.config(
            text="zone de recherche \ndu composé " + self.comps + ":"
        )
        self.tranchpotlabel.config(
            text="zone de recherche \ndu composé " + self.comppot + ":"
        )
        self.preclabel.place(x=150, y=400)
        self.tranchslabel.place(x=10, y=175)
        self.tranches1lab.place(x=self.placecurs1 + 350, y=170)
        self.tranchepot1lab.place(x=self.placecurpot1 + 350, y=270)
        self.tranches2lab.place(x=self.placecurs2 + 350, y=170)
        self.tranchepot2lab.place(x=self.placecurpot2 + 350, y=270)
        self.estimation.place(x=280, y=10)
        self.tranchpotlabel.place(x=10, y=275)
        self.situation = "parametre"
        self.tranchsliders.place(x=340, y=200)
        self.tranchsliderpot.place(x=340, y=300)
        self.precslider.place(x=340, y=400)
        self.rapportcanv1.place(x=0, y=0)
        self.preclab.place(x=630, y=390)
        self.rapportcanv12.place_forget()
        self.manuel.place_forget()
        self.experiencetexte.place_forget()
        self.experiencecurseur.place_forget()
        self.importation.place_forget()
        self.rapportconfirme1.place_forget()
        self.rapportconfirme2.place_forget()
        self.cheminex.place_forget()
        self.cheminboutonex.place_forget()
        self.cheminextext.place_forget()
        self.frame3.place_forget()
        self.label3.place_forget()
        self.composélabel.place_forget()
        self.composéAcell.place_forget()
        self.composéBcell.place_forget()

    def rapport2(self):
        self.label3.place_forget()
        self.frame3.place_forget()
        self.comps = self.composéAcell.get()
        self.comppot = self.composéBcell.get()
        self.experience = self.experiencecurseur.get()
        self.situation = "rapport2"
        self.excelcheck = 0
        self.rapportcanv1.place_forget()
        self.rapportcanv12.place_forget()
        self.importation.place_forget()
        self.manuel.place_forget()
        self.experiencetexte.place_forget()
        self.experiencecurseur.place_forget()
        self.rapportconfirme1.place_forget()
        self.cheminex.place_forget()
        self.cheminboutonex.place_forget()
        self.cheminextext.place_forget()
        self.exp = [experience(self, 1, self.comps, self.comppot)]
        self.composélabel.place_forget()
        self.composéAcell.place_forget()
        self.composéBcell.place_forget()
        self.rapportconfirme2.place_forget()
        if self.experience > len(self.exp):
            for i in range(self.experience - len(self.exp)):
                self.exp.append(
                    experience(self, len(self.exp) + 1, self.comps, self.comppot)
                )
        elif self.experience < len(self.exp):
            for i in range(len(self.exp) - self.experience):
                del self.exp[-1]
        self.exp[0].pack()
        self.expact = 0

    def expsuivante(self):
        if self.expact < len(self.exp) - 1:
            self.exp[self.expact].pack_forget()
            self.expact += 1
            self.exp[self.expact].pack()
        else:
            self.exp[self.expact].pack_forget()
            self.expact += 1

            self.chargement.place_forget()
            self.chargement = ttk.Progressbar(
                self,
                orient="horizontal",
                mode="determinate",
                maximum=100,
                value=0,
                length=400,
            )
            self.chargement.place(x=200, y=400)
            self.calcul()

    def experienceprecedente(self):
        self.exp[self.expact - 1].expabs = []
        self.exp[self.expact - 1].expord = []
        if self.expact < len(self.exp):
            self.exp[self.expact].pack_forget()
            self.expact -= 1
            self.exp[self.expact].pack()
        else:
            self.rapportcanv1.place_forget()
            self.expact -= 1
            self.exp[self.expact].pack()
            self.menubouton.place_forget()
            self.reponse.place_forget()
            self.sommecarrereponse.place_forget()
            self.widget.place_forget()
            self.frame2.place_forget()
            self.label2.pack_forget()
            self.curseurerreur.place_forget()
            self.ligne.place_forget()
            self.curseurlabel.place_forget()
            self.canvsliderpot.place_forget()
            self.canvsliders.place_forget()
            self.curseurlabel2.place_forget()

    def rapportexcel(self):
        try:
            self.docexcel = openpyxl.load_workbook(self.cheminex.get())
            self.comps = self.composéAcell.get()
            self.comppot = self.composéBcell.get()
            self.exp = []
            self.situation = "excel"
            self.excelcheck = 1
            for i in [4, 15, 33, 53, 72]:
                self.exp.append(experience2(self.docexcel, i))

            self.calcul()
        except openpyxl.utils.exceptions.InvalidFileException:
            Message(
                [
                    "erreur lors de la recherche du fichier",
                    "sources d'erreurs possibles:",
                    "- vous avez peut-être oublié de rentrer le chemin du fichier.",
                    "- vous avez peut-être fait une erreur de frappe dans le chemin du fichier.",
                    "- le fichier est peut-être dans le mauvais format: ",
                    " assurez vous qu'il soit sous format .xlsx",
                ],
                (self.winfo_x(), self.winfo_y()),
            )
        except FileNotFoundError:
            Message(
                [
                    "erreur lors de la recherche du fichier",
                    "sources d'erreurs possibles:",
                    "- vous avez peut-être oublié de rentrer le chemin du fichier.",
                    "- vous avez peut-être fait une erreur de frappe dans le chemin du fichier.",
                    "- le fichier est peut-être dans le mauvais format: ",
                    " assurez vous qu'il soit sous format .xlsx",
                ],
                (self.winfo_x(), self.winfo_y()),
            )
        except TypeError:
            Message(
                [
                    "erreur dans les données du fichier",
                    "Le fichier a été trouvé mais ses données ne sont pas exploitables.",
                    "sources d'erreurs possibles:",
                    "- vous vous êtes peut-être trompé de fichier.",
                    "- le fichier n'est peut-être pas adapté au logiciel.",
                    "inspirez vous du fichier excell-type.xlsx dans les dossiers de l'application",
                    "- le fichier a peut-être été modifié sans être sauvegardé.",
                    "cette erreur peut arriver lorsque le fichier est ouvert. essayez de le fermer.",
                ],
                (self.winfo_x(), self.winfo_y()),
            )

    def calcul(self):
        try:
            concin = []
            expl = []
            fpotl = []
            precis = 2
            ests = 1
            estpot = 1
            self.tranches = range(
                int(self.debtranchs * 100),
                int(self.fintranchs * 100),
                int(self.preclabval * 100),
            )
            self.tranchepot = range(
                int(self.debtranchpot * 100),
                int(self.fintranchpot * 100),
                int(self.preclabval * 100),
            )
            for i in range(len(self.exp)):
                concin.append(self.exp[i].concin)
                expl.append(self.exp[i].expabs)

                fpotl.append(self.exp[i].expord)

            self.label3.place_forget()
            self.frame3.place_forget()
            self.composéBcell.place_forget()
            self.composélabel.place_forget()
            self.composéAcell.place_forget()
            self.rapportcanv12.place_forget()
            self.importation.place_forget()
            self.manuel.place_forget()
            self.experiencetexte.place_forget()
            self.rapportcanv1.place_forget()
            self.experiencecurseur.place_forget()
            self.rapportconfirme1.place_forget()
            self.cheminex.place_forget()
            self.cheminboutonex.place_forget()
            self.cheminextext.place_forget()
            self.canv3 = tk.Canvas(
                self,
                bg="light green",
                height=500,
                width=800,
                bd=0,
                highlightthickness=0,
                relief="ridge",
            )
            self.chargement.place_forget()
            self.chargement = ttk.Progressbar(
                self,
                orient="horizontal",
                mode="determinate",
                maximum=100,
                value=0,
                length=400,
            )
            self.chargement.place(x=200, y=400)

            self.canv3.pack()
            self.update_idletasks()
            self.rapportconfirme2.place_forget()

            result = moindre_carré(
                self, concin, expl, fpotl, self.tranches, self.tranchepot
            )
            self.represent(result[0][0], result[0][1], result[1])
        except ZeroDivisionError:
            event = tk.Event()
            event.x = 40
            self.sommecarrereponse = tk.Label(self)
            self.widget = tk.Label(self)
            self.reponse = tk.Label(self)
            self.curseurerreur = tk.Label(self)
            self.curseurlabel = tk.Label(self)
            self.canvsliderpot = tk.Canvas(self)
            self.canvsliders = tk.Canvas(self)
            self.curseurlabel2 = tk.Label(self)
            self.ligne = tk.Canvas(self)
            self.chargement.place_forget()
            self.char.place_forget()
            self.temps.place_forget()
            event.y = 40
            self.clickdroit(mclick=event)
            Message(
                [
                    "erreur de calcul",
                    "sources d'erreurs possibles:",
                    "- le calcul avec ces données entraîne une division par 0",
                ],
                (self.winfo_x(), self.winfo_y()),
            )
        except TypeError:
            event = tk.Event()
            event.x = 40
            self.sommecarrereponse = tk.Label(self)
            self.widget = tk.Label(self)
            self.reponse = tk.Label(self)
            self.curseurerreur = tk.Label(self)
            self.curseurlabel = tk.Label(self)
            self.canvsliderpot = tk.Canvas(self)
            self.canvsliders = tk.Canvas(self)
            self.curseurlabel2 = tk.Label(self)
            self.ligne = tk.Canvas(self)
            self.chargement.place_forget()
            self.char.place_forget()
            self.temps.place_forget()
            event.y = 40
            self.clickdroit(mclick=event)
            Message(
                [
                    "erreur de calcul",
                    "sources d'erreurs possibles:",
                    "- le calcul avec ces données entraîne une racine de nombre négatif",
                ],
                (self.winfo_x(), self.winfo_y()),
            )

    def represent(self, rS, rPOT, sommecarre):
        self.rs = rS
        self.rpot = rPOT
        alpha = rS / (1 - rS)
        beta = rPOT / (1 - rPOT)
        gamma = (1 - rS * rPOT) / ((1 - rS) * (1 - rPOT))
        delta = (1 - rS) / (2 - rS - rPOT)

        self.reponse = tk.Label(
            self,
            text="r"
            + self.comps
            + " = "
            + str(rS)
            + "\nr"
            + self.comppot
            + " = "
            + str(rPOT),
            font="Helvetica 30 bold",
            bg="light green",
        )
        glup = Figure(figsize=(6, 4), dpi=75)
        self.chargement.step(10)
        ax = glup.add_subplot()
        self.fig, self.bx = plt.subplots(figsize=(6, 4))
        couleur = [
            "red",
            "royalblue",
            "palegreen",
            "yellow",
            "aqua",
            "orange",
            "magenta",
            "sienna",
            "silver",
            "lightpink",
        ]
        self.line = []
        self.lineb = []
        for i in range(len(self.exp)):
            self.line.append(2)
            self.lineb.append(2)
            ax.plot(
                self.exp[i].expabs,
                self.exp[i].expord,
                "o",
                color=couleur[i],
                label="expérience " + str(i),
            )
            self.bx.plot(
                self.exp[i].expabs,
                self.exp[i].expord,
                "o",
                color=couleur[i],
                label="expérience " + str(i),
            )
            fa0 = self.exp[i].concin

            fa = np.arange(min(0, fa0 - 0, 2), max(fa0 + 0.2, 1), 0.005)

            pfa = 1 - (fa / fa0) ** (alpha) * ((1 - fa) / (1 - fa0)) ** (beta) * (
                (fa0 - delta) / (fa - delta)
            ) ** (gamma)
            (self.line[i],) = ax.plot(pfa, fa, color=couleur[i])
            (self.lineb[i],) = self.bx.plot(
                pfa,
                fa,
                color=couleur[i],
            )

        self.sommecarrereponse = tk.Label(
            self,
            text="sommes des \ncarrés erreurs: \n" + str(round(sommecarre, 4)),
            font="Helvetica 30",
            bg="light green",
        )
        ax.set_xlabel("overall molar conversion")
        ax.set_ylabel("f " + self.comppot)
        self.bx.set_xlabel("overall molar conversion")
        self.bx.set_ylabel("f " + self.comppot)
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        self.bx.set_xlim(0, 1)
        self.bx.set_ylim(0, 1)

        self.canvas = FigureCanvasTkAgg(glup, master=self)  # A tk.DrawingArea.
        self.canvas.draw()
        self.canvas.mpl_connect(
            "key_press_event", lambda event: print(f"you pressed {event.key}")
        )
        self.canvas.mpl_connect("key_press_event", key_press_handler)
        self.widget = self.canvas.get_tk_widget()
        self.widget.postscript(file="testmatplot.ps", x=0, y=0, height=300, width=450)

        self.canvsliders = tk.Canvas(
            self,
            bg="white",
            height=20,
            width=240,
            bd=0,
            highlightthickness=2,
            relief="ridge",
        )
        self.curseurs = self.canvsliders.create_rectangle(
            (rPOT - self.debtranchpot)
            * 50
            * 4
            / (self.fintranchpot - self.debtranchpot),
            0,
            (rPOT - self.debtranchpot)
            * 50
            * 4
            / (self.fintranchpot - self.debtranchpot)
            + 40,
            20,
            fill="light blue",
            outline="grey",
        )

        self.canvsliders.bind("<B1-Motion>", self.updatecurseurs)
        self.canvsliderpot = tk.Canvas(
            self,
            bg="white",
            height=20,
            width=240,
            bd=0,
            highlightthickness=2,
            relief="ridge",
        )
        self.ligne = tk.Canvas(
            self,
            bg="black",
            height=3,
            width=800,
            bd=0,
            highlightthickness=0,
            relief="ridge",
        )

        self.curseurpot = self.canvsliderpot.create_rectangle(
            (rS - self.debtranchs) * 50 * 4 / (self.fintranchs - self.debtranchs),
            0,
            (rS - self.debtranchs) * 50 * 4 / (self.fintranchs - self.debtranchs) + 40,
            20,
            fill="light blue",
            outline="grey",
        )
        self.curseurlabel = tk.Label(
            self,
            text="r" + self.comps + " = " + str(rS),
            font="Helvetica 20 ",
            bg="light green",
        )
        self.curseurerreur = tk.Label(
            self,
            text="somme des \nerreurs carrés:\n" + str(round(sommecarre, 3)),
            font="Helvetica 18 ",
            bg="light green",
        )
        self.curseurlabel2 = tk.Label(
            self,
            text="r" + self.comppot + " = " + str(rPOT),
            font="Helvetica 20",
            bg="light green",
        )
        self.canvsliderpot.bind("<B1-Motion>", self.updatecurseurpot)
        self.affiche()

    def updatecurseurs(self, mclick):
        if mclick.x < 20:
            self.canvsliders.moveto(self.curseurs, 0, 0)
            f = self.debtranchpot
        elif mclick.x > 220:
            self.canvsliders.moveto(self.curseurs, 200, 0)
            f = self.fintranchpot
        else:
            self.canvsliders.moveto(self.curseurs, mclick.x - 20, 0)
            f = self.debtranchpot + float(mclick.x - 20) * (
                self.fintranchpot - self.debtranchpot
            ) / (50 * 4)
        self.rpot = round(f, 2)
        alpha = self.rs / (1 - self.rs)
        beta = f / (1 - f)
        gamma = (1 - self.rs * f) / ((1 - self.rs) * (1 - f))
        delta = (1 - self.rs) / (2 - self.rs - f)
        sommecarre = 0.0
        for i in range(len(self.exp)):
            fa0 = self.exp[i].concin
            if isinstance(sommecarre, float):
                try:
                    for j in range(len(self.exp[i].expabs)):
                        carre = (
                            fonction(fa0, (self.rs, self.rpot), self.exp[i].expord[j])
                            - self.exp[i].expabs[j]
                        )
                        sommecarre += carre**2
                except ZeroDivisionError:
                    sommecarre = "division par 0"
                except OverflowError:
                    sommecarre = "nombre trop grand"
            else:
                pass
            fa = np.arange(min(0, fa0 - 0, 2), max(fa0 + 0.2, 1), 0.005)

            y = 1 - (fa / fa0) ** (alpha) * ((1 - fa) / (1 - fa0)) ** (beta) * (
                (fa0 - delta) / (fa - delta)
            ) ** (gamma)
            self.line[i].set_data(y, fa)
            self.lineb[i].set_data(y, fa)

        if isinstance(sommecarre, float):
            sommecarre = round(sommecarre, 2)
        elif isinstance(sommecarre, complex):
            sommecarre = "résultat complexe"
        self.curseurlabel.configure(text="r" + self.comps + " = " + str(self.rs))

        self.curseurerreur.configure(
            text="somme des \nerreurs carrés:\n" + str(sommecarre)
        )

        self.curseurlabel2.configure(text="r" + self.comppot + " = " + str(self.rpot))
        self.canvas.draw()

    def updatecurseurpot(self, mclick):
        if mclick.x < 20:
            self.canvsliderpot.moveto(self.curseurpot, 0, 0)
            f = self.debtranchs
        elif mclick.x > 220:
            self.canvsliderpot.moveto(self.curseurpot, 200, 0)
            f = self.fintranchs
        else:
            self.canvsliderpot.moveto(self.curseurpot, mclick.x - 20, 0)
            f = self.debtranchs + float(mclick.x - 20) * (
                self.fintranchs - self.debtranchs
            ) / (50 * 4)

        self.rs = round(f, 2)
        alpha = f / (1 - f)
        beta = self.rpot / (1 - self.rpot)
        gamma = (1 - f * self.rpot) / ((1 - f) * (1 - self.rpot))
        delta = (1 - f) / (2 - f - self.rpot)
        sommecarre = 0.0
        for i in range(len(self.exp)):
            fa0 = self.exp[i].concin
            if isinstance(sommecarre, float):
                try:
                    for j in range(len(self.exp[i].expabs)):
                        carre = (
                            fonction(fa0, (self.rs, self.rpot), self.exp[i].expord[j])
                            - self.exp[i].expabs[j]
                        )
                        sommecarre += carre**2
                except ZeroDivisionError:
                    sommecarre = "division par 0"
                except OverflowError:
                    sommecarre = "nombre trop grand"
            else:
                pass
            fa = np.arange(min(0, fa0 - 0, 2), max(fa0 + 0.2, 1), 0.005)

            y = 1 - (fa / fa0) ** (alpha) * ((1 - fa) / (1 - fa0)) ** (beta) * (
                (fa0 - delta) / (fa - delta)
            ) ** (gamma)
            self.line[i].set_data(y, fa)
            self.lineb[i].set_data(y, fa)
        self.curseurlabel.configure(text="r" + self.comps + " = " + str(self.rs))
        if isinstance(sommecarre, float):
            sommecarre = round(sommecarre, 2)
            if sommecarre == 0:
                sommecarre = "erreur"
        elif isinstance(sommecarre, complex):
            sommecarre = "résultat complexe"

        self.curseurerreur.configure(
            text="somme des \nerreurs carrés:\n" + str(sommecarre)
        )
        self.curseurlabel2.configure(text="r" + self.comppot + " = " + str(self.rpot))
        self.canvas.draw()

    def affiche(self):
        self.chargement.place_forget()
        self.canv3.pack_forget()
        self.ligne.place(x=0, y=402)
        self.canvsliders.place(x=280, y=446)
        self.curseurlabel.place(x=10, y=406)
        self.curseurerreur.place(x=550, y=406)
        self.curseurlabel2.place(x=280, y=406)
        self.canvsliderpot.place(x=10, y=446)
        if self.excelcheck == 0:
            self.situation = "rapport2"
        else:
            self.situation = "excel"
        self.sommecarrereponse.place(x=10, y=250)
        self.reponse.place(x=20, y=150)
        self.rapportcanv1.place(x=0, y=0)
        self.frame2.place(x=450, y=10)
        self.label2.pack()
        self.menubouton.place(x=120, y=20)
        self.widget.place(x=320, y=100)
        self.menuderoul.place_forget()
        self.format.place_forget()
        self.chemintext.place_forget()
        self.chemin.place_forget()
        self.cheminbouton.place_forget()
        self.nom.place_forget()
        self.nomtabl.place_forget()
        self.enregistrebouton.place_forget()
        if self.enrv == 1:
            self.enrv = 0
            fini = Message(
                "le document est enregistré", (self.winfo_x(), self.winfo_y())
            )

    def save(self):
        try:
            if self.situation == "telecharge1":
                fig = self.fig
            if self.situation == "telechargeanal":
                fig2 = self.figa
                fig1 = self.figa2
            if self.situation == "telechargemc":
                try:
                    fig1 = self.figmc
                except AttributeError:
                    fig1 = "hit the road"
                figd = Image.open("simulmc.bmp")
                figdo = Image.open("simulmc2.bmp")
            chemin = self.chemin.get()

            nom = self.nomtabl.get()
            format = self.menuderoul.get()
            if format == "tableur (format xlsx)":
                shutil.copyfile(
                    "excel-type.xlsx",
                    chemin + "/" + nom + ".xlsx",
                )
                file = openpyxl.load_workbook(chemin + "/" + nom + ".xlsx")
                sheet = file.active
                emplace = [4, 15, 33, 53, 72]
                sheet["A2"] = self.comps + " = A"
                sheet["A3"] = self.comppot + " = B"
                sheet["E2"] = self.rs
                sheet["E3"] = self.rpot
                for i in range(min(len(self.exp), 5)):
                    sheet["L" + str(emplace[i])] = round(self.exp[i].concin, 5)
                    for j in range(len(self.exp[i].expabs)):  # x expt
                        sheet["O" + str(emplace[i] + j)] = round(
                            self.exp[i].expabs[j], 5
                        )
                    for j in range(len(self.exp[i].expord)):  # fpot expt
                        sheet["P" + str(emplace[i] + j)] = round(
                            self.exp[i].expord[j], 5
                        )

                file.save(chemin + "/" + nom + ".xlsx")
                self.enrv = 1
            elif format == "image (format png)":
                fig.savefig(chemin + "/" + nom + ".png")
                self.enrv = 1
            elif format == "image (format pdf)":
                fig.savefig(chemin + "/" + nom + ".pdf")
                self.enrv = 1
            elif format == "image (format jpg)":
                fig.savefig(chemin + "/" + nom + ".jpg")
                self.enrv = 1
            elif format == "graphique 2D (format png)":
                fig1.savefig(chemin + "/" + nom + ".png")
                self.enrv = 1
            elif format == "graphique 2D (format pdf)":
                fig1.savefig(chemin + "/" + nom + ".pdf")
                self.enrv = 1
            elif format == "graphique 2D (format jpg)":
                fig1.savefig(chemin + "/" + nom + ".jpg")
                self.enrv = 1
            elif format == "graphique (format png)":
                if isinstance(fig1, str):
                    Message(
                        [
                            "erreur lors de la genération de l'image",
                            "appuyez sur le bouton 'distribution' dans le menu précédent puis réessayer",
                        ],
                        (self.winfo_x(), self.winfo_y()),
                    )
                else:
                    fig1.savefig(chemin + "/" + nom + ".png")
                    self.enrv = 1
            elif format == "graphique (format pdf)":
                if isinstance(fig1, str):
                    Message(
                        [
                            "erreur lors de la genération de l'image",
                            "appuyez sur le bouton 'distribution' dans le menu précédent puis réessayer",
                        ],
                        (self.winfo_x(), self.winfo_y()),
                    )
                else:
                    fig1.savefig(chemin + "/" + nom + ".pdf")
                    self.enrv = 1
            elif format == "graphique (format jpg)":
                if isinstance(fig1, str):
                    Message(
                        [
                            "erreur lors de la genération de l'image",
                            "appuyez sur le bouton 'distribution' dans le menu précédent puis réessayer",
                        ],
                        (self.winfo_x(), self.winfo_y()),
                    )
                else:
                    fig1.savefig(chemin + "/" + nom + ".jpg")
                    self.enrv = 1
            elif format == "graphique 3D (format png)":
                fig2.savefig(chemin + "/" + nom + ".png")
                self.enrv = 1
            elif format == "graphique 3D (format pdf)":
                fig2.savefig(chemin + "/" + nom + ".pdf")
                self.enrv = 1
            elif format == "graphique 3D (format jpg)":
                fig2.savefig(chemin + "/" + nom + ".jpg")
                self.enrv = 1
            elif format == "distribution (format png)":
                figd.save(chemin + "/" + nom + ".png")
                self.enrv = 1
            elif format == "distribution (format pdf)":
                figd.save(chemin + "/" + nom + ".pdf")
                self.enrv = 1
            elif format == "distribution (format jpg)":
                figd.save(chemin + "/" + nom + ".jpg")
                self.enrv = 1
            elif format == "distribution ordoné (format png)":
                figdo.save(chemin + "/" + nom + ".png")
                self.enrv = 1
            elif format == "distribution ordoné (format pdf)":
                figdo.save(chemin + "/" + nom + ".pdf")
                self.enrv = 1
            elif format == "distribution ordoné (format jpg)":
                figdo.save(chemin + "/" + nom + ".jpg")
                self.enrv = 1
            if self.situation == "telecharge1":
                self.affiche()
            if self.situation == "telechargeanal":
                self.simuldist()
            if self.situation == "telechargemc":
                self.affichemc()

        except openpyxl.utils.exceptions.InvalidFileException:
            Message(
                [
                    "erreur lors de l'enregistrement du fichier",
                    "sources d'erreurs possibles:",
                    "- vous avez peut-être oublier de rentrer le chemin ou le nom du fichier.",
                    "- vous avez peut-être fait une erreur de frappe dans le chemin du fichier.",
                    "-le fichier excel-type des dossiers du programme a peut-être été compromis",
                ],
                (self.winfo_x(), self.winfo_y()),
            )
        except FileNotFoundError:
            Message(
                [
                    "erreur lors de l'enregistrement du fichier",
                    "sources d'erreurs possibles:",
                    "- vous avez peut-être oublier de rentrer le chemin ou le nom du fichier.",
                    "- vous avez peut-être fait une erreur de frappe dans le chemin du fichier.",
                    "-le fichier excel-type des dossiers du programme a peut-être été compromis",
                ],
                (self.winfo_x(), self.winfo_y()),
            )
        except UnboundLocalError:
            Message(
                [
                    "erreur lors de l'enregistrement du fichier",
                    "sources d'erreurs possibles:",
                    "- vous avez peut-être oublier de rentrer le chemin ou le nom du fichier.",
                    "- vous avez peut-être fait une erreur de frappe dans le chemin du fichier.",
                    "-le fichier excel-type des dossiers du programme a peut-être été compromis",
                ],
                (self.winfo_x(), self.winfo_y()),
            )

    def fini(self):
        self.curseurerreur.place_forget()
        self.canvsliders.place_forget()
        self.curseurlabel.place_forget()
        self.curseurlabel2.place_forget()
        self.canvsliderpot.place_forget()
        self.ligne.place_forget()
        self.frame2.place_forget()
        self.label2.pack_forget()
        self.situation = "menu"
        self.exp = [experience(self, 1, self.comps, self.comppot)]
        self.canvmenu.pack()
        self.rapportbout.pack()
        self.analbout.place(x=400, y=325, anchor=tk.CENTER)
        self.mcbout.place(x=400, y=395, anchor=tk.CENTER)
        self.rapportcanv1.place_forget()
        self.menubouton.place_forget()
        self.reponse.place_forget()
        self.sommecarrereponse.place_forget()
        self.frame.place(x=100, y=10)
        self.label.pack()
        self.widget.place_forget()

    def enregistre(self, mclick):
        if self.situation == "excel":
            self.situation = "telecharge1"

            self.frame2.place_forget()
            self.label2.pack_forget()
            self.menubouton.place_forget()
            self.reponse.place_forget()
            self.sommecarrereponse.place_forget()
            self.widget.place_forget()
            self.curseurerreur.place_forget()
            self.ligne.place_forget()
            self.canvsliders.place_forget()
            self.curseurlabel.place_forget()
            self.curseurlabel2.place_forget()
            self.canvsliderpot.place_forget()
            self.menuderoul.config(
                values=[
                    "tableur (format xlsx)",
                    "image (format png)",
                    "image (format jpg)",
                    "image (format pdf)",
                ]
            )
        if self.situation == "rapport2":
            self.situation = "telecharge1"
            self.frame2.place_forget()
            self.label2.pack_forget()
            self.menubouton.place_forget()
            self.reponse.place_forget()
            self.sommecarrereponse.place_forget()
            self.widget.place_forget()
            self.curseurerreur.place_forget()
            self.ligne.place_forget()
            self.canvsliders.place_forget()
            self.curseurlabel.place_forget()
            self.curseurlabel2.place_forget()
            self.canvsliderpot.place_forget()
            self.menuderoul.config(
                values=[
                    "tableur (format xlsx)",
                    "image (format png)",
                    "image (format jpg)",
                    "image (format pdf)",
                ]
            )
        elif self.situation == "simuldistrib":
            self.situation = "telechargeanal"
            self.widget2.place_forget()
            self.widget3.place_forget()
            self.radio1.place_forget()
            self.frameanal1.place_forget()
            self.labelanal1.pack_forget()
            self.radio2.place_forget()
            self.cursor.place_forget()
            self.menuderoul.config(
                values=[
                    "graphique 2D (format png)",
                    "graphique 2D (format jpg)",
                    "graphique 2D (format pdf)",
                    "graphique 3D (format png)",
                    "graphique 3D (format jpg)",
                    "graphique 3D (format pdf)",
                ]
            )
        elif self.situation == "affichemc":
            self.situation = "telechargemc"
            self.simulmc.place_forget()
            self.simulmc2.place_forget()
            self.framemc.place_forget()
            self.labelmc.pack_forget()
            self.distribmc.place_forget()
            self.radiomc.place_forget()
            self.monoremc.place_forget()
            for i in self.monolistlabel:
                i.place_forget()
            try:
                self.widget4.place_forget()
            except AttributeError:
                pass
            self.menuderoul.config(
                values=[
                    "graphique (format png)",
                    "graphique (format jpg)",
                    "graphique (format pdf)",
                    "distribution (format png)",
                    "distribution (format jpg)",
                    "distribution (format pdf)",
                    "distribution ordoné (format png)",
                    "distribution ordoné (format jpg)",
                    "distribution ordoné (format pdf)",
                ]
            )
        self.menuderoul.place(x=300, y=200)
        self.format.place(x=100, y=200)
        self.chemintext.place(x=30, y=315)
        self.chemin.place(x=240, y=320)
        self.cheminbouton.place(x=400, y=360)
        self.nom.place(x=140, y=255)
        self.nomtabl.place(x=240, y=260)
        self.enregistrebouton.place(x=200, y=420)


myapp = aplli()
myapp.title("polyphème")
myapp.mainloop()
